"""
=========================================================
KiG POS
=========================================================

Datei:
    screens/cash_book_screen.py

Beschreibung:
    Kassenbuch: Was liegt tatsächlich in der Kasse?

    Je Tag eine Zeile - Startbestand, Einnahmen, Ausgaben,
    Endbestand, dazu ein Kommentar und der Name dessen, der
    nachgezählt hat.

    Bewusst getrennt von der Statistik: Die beantwortet, was
    verkauft wurde. Hier geht es um den Kassenbestand, in den
    auch Einlagen, Entnahmen und Wechselgeld einfließen.

    Aufbau (Querformat):

        links    Jahr, darunter Monat
        Mitte    Übersichtstabelle des gewählten Monats
        rechts   Eingabefeld für eine Zeile

    Geht die Rechnung einer Zeile nicht auf, steht in der
    Spalte "Kommentar" ein rotes "Prüfen" - siehe
    database.py:cash_book_entry_is_valid.

Version:
    1.0.0
=========================================================
"""

from datetime import date, datetime

from kivy.clock import Clock
from kivy.graphics import Color, RoundedRectangle
from kivy.metrics import dp
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.screenmanager import Screen
from kivy.uix.widget import Widget

import config
import geldformat
import storage
import teilen
import theme

from database import DatabaseManager
from widgets.common.exporthinweis import (
    export_hinweis, hinweisfeld_vorbereiten,
)
from widgets.common.feldausrichtung import links_ausrichten
from widgets.common.confirm_popup import ConfirmPopup
from widgets.common.date_picker_popup import DatePickerPopup
from widgets.common.rounded_input import RoundedInput
from widgets.common.rounded_panel import RoundedPanel
from widgets.kig_label import KiGLabel


MONTH_NAMES = (
    "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember"
)


# Spaltenaufteilung der Tabelle. Kopfzeile und Datenzeilen benutzen
# dieselbe Liste - nur gemeinsam ändern, sonst stehen Überschrift und
# Wert nicht mehr übereinander.
COLUMNS = (
    ("Datum", 0.14),
    ("Startbestand", 0.15),
    ("Einnahmen", 0.14),
    ("Ausgaben", 0.14),
    ("Endbestand", 0.15),
    ("Kommentar", 0.16),
    ("Prüfer", 0.12),
)


# Auf dem Telefon vier statt sieben Spalten. Sieben Ueberschriften
# ergaben auf 412 dp Breite einen einzigen unlesbaren Streifen
# ("DatumStartbestaEinnahmeAusgabenEndbestaKommentaPruefer").
# Startbestand, Kommentar und Pruefer stehen weiterhin im Formular
# darunter, sobald eine Zeile angetippt wird.
NARROW_COLUMNS = (
    ("Datum", 0.28),
    ("Einnahmen", 0.24),
    ("Ausgaben", 0.24),
    ("Endbestand", 0.24),
)


def spalten():
    """Die Spalten dieses Bildschirms - je nach Breite."""

    return NARROW_COLUMNS if theme.is_narrow() else COLUMNS


class CashBookRow(ButtonBehavior, BoxLayout):
    """Eine Zeile der Übersicht. Ein Tipp lädt sie zum Bearbeiten."""

    HEIGHT = 44

    def __init__(self, entry, problems, selected_callback, **kwargs):

        super().__init__(
            orientation="horizontal", size_hint_y=None,
            height=dp(self.HEIGHT), **kwargs
        )

        self.entry = entry
        self.problems = problems or []
        self.valid = not self.problems
        self.selected_callback = selected_callback
        self.selected = False

        with self.canvas.before:
            self._color = Color(*theme.CARD)
            self._background = RoundedRectangle(
                pos=self.pos, size=self.size, radius=[dp(6)]
            )

        self.bind(pos=self._refresh_canvas, size=self._refresh_canvas)

        # Gibt es einen Befund, ersetzt "Prüfen" den Kommentar - ein
        # Hinweis, der untergeht, ist keiner.
        kommentar = "Prüfen" if self.problems else (entry["comment"] or "")

        datum = CashBookScreen.format_date(entry["entry_date"])

        if theme.is_narrow():

            # Ohne Kommentarspalte braeuchte der Befund einen anderen
            # Platz - er wandert an das Datum, damit ein Hinweis nicht
            # verlorengeht.
            werte = (
                f"! {datum}" if self.problems else datum,
                CashBookScreen.money(entry["income"]),
                CashBookScreen.money(entry["expenses"]),
                CashBookScreen.money(entry["closing_balance"]),
            )

            befund_spalte = 0

        else:

            werte = (
                datum,
                CashBookScreen.money(entry["opening_balance"]),
                CashBookScreen.money(entry["income"]),
                CashBookScreen.money(entry["expenses"]),
                CashBookScreen.money(entry["closing_balance"]),
                kommentar,
                entry["auditor"] or "",
            )

            befund_spalte = 5

        for spalte, (wert, (_titel, breite)) in enumerate(
                zip(werte, spalten())
        ):

            farbe = (
                theme.ERROR
                if (self.problems and spalte == befund_spalte)
                else theme.TEXT_PRIMARY
            )

            label = Label(
                text=str(wert), color=farbe, font_size="13sp",
                bold=bool(self.problems and spalte == befund_spalte),
                halign="left", valign="middle",
                text_size=(None, dp(self.HEIGHT)), size_hint_x=breite,
                shorten=True, shorten_from="right",
            )

            self.add_widget(label)

    def _refresh_canvas(self, *_args):
        self._background.pos = self.pos
        self._background.size = self.size

    def on_release(self):
        self.selected_callback(self)

    def select(self):
        self.selected = True
        self._color.rgba = theme.PRIMARY_ORANGE_LIGHT

    def unselect(self):
        self.selected = False
        self._color.rgba = theme.CARD


class CashBookScreen(Screen):
    """Kassenbuch mit Jahres- und Monatsauswahl."""

    # Hoehe einer Formularzeile. Auf dem Telefon flacher: Dort sind
    # von acht Zeilen sonst nur eineinhalb zu sehen.
    ZEILENHOEHE = 52

    YEAR_BUTTON_HEIGHT = 48
    NARROW_YEAR_BUTTON_HEIGHT = 40
    MONTH_BUTTON_HEIGHT = 42
    NARROW_MONTH_BUTTON_HEIGHT = 34
    SELECTION_WIDTH = 190

    def __init__(self, **kwargs):

        super().__init__(**kwargs)

        self.db = DatabaseManager()

        heute = date.today()

        self.selected_year = heute.year
        self.selected_month = heute.month

        if theme.is_narrow():
            self.ZEILENHOEHE = 46
            self.YEAR_BUTTON_HEIGHT = self.NARROW_YEAR_BUTTON_HEIGHT
            self.MONTH_BUTTON_HEIGHT = self.NARROW_MONTH_BUTTON_HEIGHT

        self.selected_entry_id = None
        self.selected_row = None

        # Zuletzt ausgegebene Datei - sie haengt am Teilen-Knopf.
        self.letzte_ausgabe = None

        self.year_buttons = {}
        self.month_buttons = {}

        # Im Hochformat steht die Eingabe unter der Tabelle statt
        # daneben - nebeneinander bliebe von beidem ein Streifen.
        self.hochformat = theme.is_portrait()

        root = BoxLayout(
            orientation="vertical" if self.hochformat else "horizontal",
            padding=dp(theme.SCREEN_PADDING),
            spacing=dp(theme.SCREEN_SPACING),
        )

        root.add_widget(self._build_selection_panel())
        root.add_widget(self._build_table_panel())
        root.add_widget(self._build_form_panel())

        self.add_widget(root)

    # =====================================================
    # Links: Jahr und Monat
    # =====================================================

    def _build_selection_panel(self):

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
            size_hint=(
                (1, 0.22 if theme.is_narrow() else 0.16)
                if self.hochformat else (None, 1)
            ),
        )

        if not self.hochformat:
            panel.width = dp(self.SELECTION_WIDTH)

        panel.add_widget(self._title("Zeitraum"))

        # Im Hochformat nebeneinander: Dort ist Höhe knapp, Breite
        # nicht.
        inhalt = BoxLayout(
            orientation="horizontal" if self.hochformat else "vertical",
            spacing=dp(theme.CARD_SPACING),
        )

        self.year_box = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.SPACE_XS),
            size_hint_y=None,
        )
        self.year_box.bind(minimum_height=self.year_box.setter("height"))

        # Die Jahre bekommen nur so viel Platz, wie sie brauchen (bis
        # zu drei auf einen Blick) - der Rest gehoert den Monaten.
        self.year_scroll = ScrollView(do_scroll_x=False, bar_width=dp(8))
        self.year_scroll.add_widget(self.year_box)

        if not self.hochformat:
            self.year_scroll.size_hint_y = None

        inhalt.add_widget(self.year_scroll)

        self.month_box = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.SPACE_XS),
            size_hint_y=None,
        )
        self.month_box.bind(minimum_height=self.month_box.setter("height"))

        self.month_scroll = ScrollView(do_scroll_x=False, bar_width=dp(8))
        self.month_scroll.add_widget(self.month_box)
        inhalt.add_widget(self.month_scroll)

        panel.add_widget(inhalt)

        return panel

    def _build_year_buttons(self):

        self.year_box.clear_widgets()
        self.year_buttons = {}

        for jahr in self.db.get_cash_book_years():

            button = self._selection_button(
                str(jahr), dp(self.YEAR_BUTTON_HEIGHT),
                lambda jahr=jahr: self.select_year(jahr),
            )

            self.year_buttons[jahr] = button
            self.year_box.add_widget(button)

        if self.year_scroll.size_hint_y is None:

            # Mindestens zwei Jahre auf einen Blick - eines allein
            # sieht nach einer Schaltflaeche aus, nicht nach einer
            # Liste, in der man blaettern kann.
            sichtbar = min(3, max(2, len(self.year_buttons)))

            self.year_scroll.height = sichtbar * (
                dp(self.YEAR_BUTTON_HEIGHT) + dp(theme.SPACE_XS)
            )

        self._highlight_selection()

    def _build_month_buttons(self):

        self.month_box.clear_widgets()
        self.month_buttons = {}

        for nummer, name in enumerate(MONTH_NAMES, start=1):

            button = self._selection_button(
                name, dp(self.MONTH_BUTTON_HEIGHT),
                lambda nummer=nummer: self.select_month(nummer),
            )

            self.month_buttons[nummer] = button
            self.month_box.add_widget(button)

        self._highlight_selection()

        # Ohne das startet die Liste im Januar - der laufende Monat
        # waere ausgerechnet dann nicht zu sehen, wenn man ihn braucht.
        Clock.schedule_once(lambda _dt: self._scroll_to_month(), 0)

    def _scroll_to_month(self):

        button = self.month_buttons.get(self.selected_month)

        if button is not None and button.parent is not None:
            self.month_scroll.scroll_to(button, padding=dp(20), animate=False)

    def _selection_button(self, text, height, callback):

        button = Button(
            text=text, size_hint_y=None, height=height,
            background_normal="", background_down="",
            background_color=theme.SURFACE, color=theme.TEXT_PRIMARY,
            font_size="15sp", bold=True,
        )

        button.bind(on_release=lambda *_args: callback())

        return button

    def _highlight_selection(self):
        """Färbt die gewählte Jahres- und Monatsschaltfläche."""

        for jahr, button in self.year_buttons.items():
            gewaehlt = jahr == self.selected_year
            button.background_color = (
                theme.PRIMARY_ORANGE if gewaehlt else theme.SURFACE
            )
            button.color = theme.TEXT_WHITE if gewaehlt else theme.TEXT_PRIMARY

        for monat, button in self.month_buttons.items():
            gewaehlt = monat == self.selected_month
            button.background_color = (
                theme.PRIMARY_ORANGE if gewaehlt else theme.SURFACE
            )
            button.color = theme.TEXT_WHITE if gewaehlt else theme.TEXT_PRIMARY

    def select_year(self, jahr):

        self.selected_year = jahr
        self._highlight_selection()
        self.refresh()

    def select_month(self, monat):

        self.selected_month = monat
        self._highlight_selection()
        self.refresh()

    # =====================================================
    # Mitte: Übersichtstabelle
    # =====================================================

    def _build_table_panel(self):

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
            size_hint=(
                (1, 0.40 if theme.is_narrow() else 0.52)
                if self.hochformat else (1, 1)
            ),
        )

        self.table_title = self._title("Kassenbuch")
        panel.add_widget(self.table_title)

        # Ausgabe zum Abheften: Der Kassenprüfer bekommt den Monat auf
        # Papier, nicht das Tablet in die Hand.
        aktionen = BoxLayout(
            size_hint_y=None, height=dp(40), spacing=dp(theme.ROW_SPACING)
        )

        aktionen.add_widget(Widget())

        schmal = theme.is_narrow()

        export_knopf = self._action_button(
            "Export" if schmal else "Excel exportieren", self.export_excel
        )
        teilen_knopf = self._action_button("Teilen", self.teilen_clicked)

        if schmal:
            # 190 + 110 dp passen auf ein Telefon nicht neben den
            # Platzhalter - dort teilen sich beide, was da ist.
            aktionen.clear_widgets()
            aktionen.add_widget(export_knopf)
            aktionen.add_widget(teilen_knopf)

        else:

            export_knopf.size_hint_x = None
            export_knopf.width = dp(190)
            aktionen.add_widget(export_knopf)

            teilen_knopf.size_hint_x = None
            teilen_knopf.width = dp(110)
            aktionen.add_widget(teilen_knopf)

        panel.add_widget(aktionen)

        # Der Hinweis steht unter der Zeile und nicht daneben: Er nennt
        # den Ordner mit, und ein vollständiger Pfad braucht die ganze
        # Breite (siehe widgets/common/exporthinweis.py).
        self.export_status = Label(
            text="", color=theme.TEXT_SECONDARY, font_size="13sp",
            halign="left", valign="middle",
        )

        hinweisfeld_vorbereiten(self.export_status, 0)

        panel.add_widget(self.export_status)

        header = BoxLayout(size_hint_y=None, height=dp(32), spacing=0)

        for titel, breite in spalten():
            header.add_widget(Label(
                text=titel, bold=True, color=theme.TEXT_PRIMARY,
                font_size="13sp", halign="left", valign="middle",
                text_size=(None, dp(32)), size_hint_x=breite,
            ))

        panel.add_widget(header)

        self.rows_box = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.SPACE_XS),
            size_hint_y=None,
        )
        self.rows_box.bind(minimum_height=self.rows_box.setter("height"))

        scroll = ScrollView(do_scroll_x=False)
        scroll.add_widget(self.rows_box)
        panel.add_widget(scroll)

        self.summary_label = Label(
            text="", color=theme.TEXT_SECONDARY, font_size="14sp",
            size_hint_y=None, height=dp(46),
            halign="left", valign="middle",
        )
        self.summary_label.bind(
            size=lambda instance, value: setattr(instance, "text_size", value)
        )
        panel.add_widget(self.summary_label)

        return panel

    # =====================================================
    # Rechts: Eingabe
    # =====================================================

    def _build_form_panel(self):

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
            size_hint=(
                (1, 0.38 if theme.is_narrow() else 0.32)
                if self.hochformat else (None, 1)
            ),
        )

        if not self.hochformat:
            panel.width = dp(theme.CART_PANEL_WIDTH)

        self.form_title = self._title("Neue Zeile")
        panel.add_widget(self.form_title)

        # Acht Zeilen zu je gut 50 dp passen auf einem Telefon nicht
        # neben die Tabelle - und im Hochformat erst recht nicht
        # darunter. Die Felder bekommen deshalb einen Rollbereich;
        # Überschrift und Schaltflächen bleiben stehen, damit
        # "Speichern" immer erreichbar ist.
        self.form_fields = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.CARD_SPACING),
            size_hint_y=None,
        )
        self.form_fields.bind(
            minimum_height=self.form_fields.setter("height")
        )

        form_scroll = ScrollView(do_scroll_x=False, bar_width=dp(8))
        form_scroll.add_widget(self.form_fields)
        panel.add_widget(form_scroll)

        zeile, self.date_button = self._form_button(
            "Datum", "", lambda: self.open_date_picker()
        )
        self.form_fields.add_widget(zeile)

        self.amount_buttons = {}

        for schluessel, beschriftung in (
            ("opening_balance", "Startbestand"),
            ("income", "Einnahmen"),
            ("expenses", "Ausgaben"),
            ("closing_balance", "Endbestand"),
        ):
            zeile, button = self._form_button(
                beschriftung, self.money(0),
                lambda schluessel=schluessel: self.open_amount_numpad(schluessel),
            )
            self.amount_buttons[schluessel] = button
            self.form_fields.add_widget(zeile)

        zeile, self.comment_input = self._form_input(
            "Kommentar", "z. B. Stadtfest"
        )
        self.form_fields.add_widget(zeile)

        zeile, self.auditor_input = self._form_input("Prüfer", "Name")
        self.form_fields.add_widget(zeile)

        # Zeigt beim Tippen mit, ob die Rechnung aufgeht - lieber
        # gleich als erst in der Tabelle.
        self.check_label = Label(
            text="", color=theme.TEXT_SECONDARY, font_size="13sp",
            size_hint_y=None, height=dp(34),
            halign="left", valign="middle",
        )
        self.check_label.bind(
            size=lambda instance, value: setattr(instance, "text_size", value)
        )
        self.form_fields.add_widget(self.check_label)

        buttons = BoxLayout(
            size_hint_y=None, height=dp(52), spacing=dp(theme.ROW_SPACING)
        )
        buttons.add_widget(self._action_button("Löschen", self.delete_entry))
        buttons.add_widget(self._action_button(
            "Speichern", self.save_entry,
            background=theme.PRIMARY_ORANGE, color=theme.TEXT_WHITE,
        ))
        panel.add_widget(buttons)

        return panel

    def _form_button(self, beschriftung, wert, callback):
        """Beschriftung links, Wert als Schaltfläche rechts.

        Liefert (Zeile, Schaltfläche): Die Zeile kommt ins Panel, die
        Schaltfläche behält der Screen, um ihren Wert zu lesen und zu
        setzen.
        """

        hoehe = dp(self.ZEILENHOEHE)

        zeile = BoxLayout(
            size_hint_y=None, height=hoehe, spacing=dp(theme.ROW_SPACING)
        )

        zeile.add_widget(Label(
            text=beschriftung, color=theme.TEXT_SECONDARY, font_size="14sp",
            halign="left", valign="middle",
            size_hint_x=0.34 if theme.is_narrow() else 0.42,
            text_size=(None, hoehe),
        ))

        button = Button(
            text=wert, background_normal="", background_down="",
            background_color=theme.SURFACE, color=theme.TEXT_PRIMARY,
            font_size="16sp", bold=True,
            size_hint_x=0.66 if theme.is_narrow() else 0.58,
        )

        # Hinter der Schaltflaeche steckt ein Kalender oder der
        # Nummernblock - sie ist ein Feld und schreibt deshalb links.
        links_ausrichten(button)

        button.bind(on_release=lambda *_args: callback())

        zeile.add_widget(button)

        return zeile, button

    def _form_input(self, beschriftung, hinweis):

        hoehe = dp(self.ZEILENHOEHE)

        zeile = BoxLayout(
            size_hint_y=None, height=hoehe, spacing=dp(theme.ROW_SPACING)
        )

        zeile.add_widget(Label(
            text=beschriftung, color=theme.TEXT_SECONDARY, font_size="14sp",
            halign="left", valign="middle",
            size_hint_x=0.34 if theme.is_narrow() else 0.42,
            text_size=(None, hoehe),
        ))

        feld = RoundedInput(
            hint_text=hinweis, multiline=False,
            size_hint_x=0.66 if theme.is_narrow() else 0.58,
        )
        feld.foreground_color = theme.INPUT_TEXT
        feld.hint_text_color = theme.INPUT_HINT

        zeile.add_widget(feld)

        return zeile, feld

    @staticmethod
    def _action_button(text, callback, background=None, color=None):

        button = Button(
            text=text, background_normal="", background_down="",
            background_color=background or theme.SURFACE,
            color=color or theme.TEXT_PRIMARY,
            font_size="15sp", bold=True,
        )
        button.bind(on_release=lambda *_args: callback())

        return button

    @staticmethod
    def _title(text):

        schmal = theme.is_narrow()

        label = KiGLabel(text=text)
        label.set_font_size(18 if schmal else 24)
        label.set_bold(True)
        label.set_alignment("left")
        label.set_color(theme.PRIMARY_ORANGE)
        label.size_hint_y = None
        label.height = dp(28 if schmal else 36)

        return label

    # =====================================================
    # Formatierung
    # =====================================================

    @staticmethod
    def money(value):
        return geldformat.geld(value)


    def teilen_clicked(self):
        """Gibt die zuletzt ausgegebene Datei weiter (siehe teilen.py)."""

        erfolg, meldung = teilen.teilen(self.letzte_ausgabe)

        self.export_status.text = meldung

    @staticmethod
    def format_date(iso_date):
        try:
            return datetime.strptime(iso_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        except (TypeError, ValueError):
            return str(iso_date or "-")

    # =====================================================
    # Screen wird geöffnet
    # =====================================================

    def on_pre_enter(self, *_args):

        self._build_year_buttons()
        self._build_month_buttons()

        self.new_entry()
        self.refresh()

    # =====================================================
    # Tabelle füllen
    # =====================================================

    def refresh(self):

        self.rows_box.clear_widgets()
        self.selected_row = None

        eintraege, befunde = self.db.get_cash_book_entries_checked(
            self.selected_year, self.selected_month
        )

        self.table_title.text = (
            f"Kassenbuch {MONTH_NAMES[self.selected_month - 1]} "
            f"{self.selected_year}"
        )

        if not eintraege:

            self.rows_box.add_widget(Label(
                text="Für diesen Monat ist noch nichts erfasst.",
                color=theme.TEXT_SECONDARY, font_size="14sp",
                size_hint_y=None, height=dp(42),
                halign="left", valign="middle", text_size=(None, dp(42)),
            ))

        else:

            for eintrag in eintraege:

                zeile = CashBookRow(
                    entry=eintrag,
                    problems=befunde.get(eintrag["id"]),
                    selected_callback=self.row_selected,
                )

                self.rows_box.add_widget(zeile)

                if eintrag["id"] == self.selected_entry_id:
                    zeile.select()
                    self.selected_row = zeile

        self._refresh_summary()

    def _refresh_summary(self):

        summen = self.db.get_cash_book_totals(
            self.selected_year, self.selected_month
        )

        if not summen["entries"]:
            self.summary_label.text = ""
            return

        text = (
            f"{summen['entries']} Tage   ·   "
            f"Einnahmen {self.money(summen['income'])}   ·   "
            f"Ausgaben {self.money(summen['expenses'])}   ·   "
            f"Kassenstand zuletzt {self.money(summen['closing_balance'])}"
        )

        if summen["invalid"]:
            offen = summen["invalid"]
            text += (
                f"   ·   {offen} {'Zeile' if offen == 1 else 'Zeilen'} "
                "zu prüfen"
            )

        self.summary_label.text = text

        self.summary_label.color = (
            theme.ERROR if summen["invalid"] else theme.TEXT_SECONDARY
        )

    # =====================================================
    # Zeile gewählt
    # =====================================================

    def row_selected(self, row):
        """Ein Tipp lädt die Zeile ins Eingabefeld, ein zweiter hebt
        die Auswahl wieder auf."""

        if self.selected_row is row:
            row.unselect()
            self.selected_row = None
            self.new_entry()
            return

        if self.selected_row is not None:
            self.selected_row.unselect()

        row.select()
        self.selected_row = row

        self.load_entry(row.entry)

    def load_entry(self, entry):

        self.selected_entry_id = entry["id"]

        self.form_title.text = "Zeile bearbeiten"

        self.date_value = entry["entry_date"]
        self.date_button.text = self.format_date(entry["entry_date"])

        for schluessel, button in self.amount_buttons.items():
            button.text = self.money(entry[schluessel])

        self.comment_input.text = entry["comment"] or ""
        self.auditor_input.text = entry["auditor"] or ""

        self._refresh_check()

    # =====================================================
    # Eingabe
    # =====================================================

    def new_entry(self):
        """Leert das Eingabefeld für eine neue Zeile.

        Vorbelegt wird der heutige Tag (bzw. der Monatsanfang, wenn ein
        anderer Monat gewählt ist) und als Startbestand der Endbestand
        des letzten Eintrags davor - in der Kasse liegt am Morgen das,
        was am Abend zuvor drin lag.
        """

        if self.selected_row is not None:
            self.selected_row.unselect()
            self.selected_row = None

        self.selected_entry_id = None
        self.form_title.text = "Neue Zeile"

        heute = date.today()

        if (heute.year, heute.month) == (self.selected_year, self.selected_month):
            vorschlag = heute
        else:
            vorschlag = date(self.selected_year, self.selected_month, 1)

        self.date_value = vorschlag.isoformat()
        self.date_button.text = self.format_date(self.date_value)

        vorheriger = self.db.get_previous_closing_balance(self.date_value)

        for schluessel, button in self.amount_buttons.items():
            button.text = self.money(0)

        if vorheriger:
            self.amount_buttons["opening_balance"].text = self.money(vorheriger)
            self.amount_buttons["closing_balance"].text = self.money(vorheriger)

        self.comment_input.text = ""
        self.auditor_input.text = ""

        self._refresh_check()

    def open_date_picker(self):

        DatePickerPopup(
            title="Datum der Kassenbuchzeile",
            initial_date=self.date_value,
            on_select=self.date_picked,
        ).open()

    def date_picked(self, iso_date):

        self.date_value = iso_date
        self.date_button.text = self.format_date(iso_date)

        # Zum Datum passt ein anderer Vorgänger - aber nur, solange
        # eine neue Zeile erfasst wird.
        if self.selected_entry_id is None:

            vorheriger = self.db.get_previous_closing_balance(iso_date)

            if vorheriger:
                self.amount_buttons["opening_balance"].text = self.money(vorheriger)

        self._refresh_check()

    def open_amount_numpad(self, schluessel):

        from widgets.common.numpad.numpad_popup import NumpadPopup

        beschriftungen = {
            "opening_balance": "Startbestand",
            "income": "Einnahmen",
            "expenses": "Ausgaben",
            "closing_balance": "Endbestand",
        }

        NumpadPopup(
            title=beschriftungen[schluessel],
            value=int(round(self._amount(schluessel) * 100)),
            mode="price",
            on_confirm=lambda cent: self.amount_entered(schluessel, cent),
        ).open()

    def amount_entered(self, schluessel, cent):

        self.amount_buttons[schluessel].text = self.money(cent / 100)

        # Endbestand mitrechnen, solange er noch nicht von Hand
        # angefasst wurde: Start + Einnahmen - Ausgaben ist der
        # Normalfall, alles andere die Ausnahme.
        if schluessel != "closing_balance":
            self.amount_buttons["closing_balance"].text = self.money(
                self._amount("opening_balance")
                + self._amount("income")
                - self._amount("expenses")
            )

        self._refresh_check()

    def _amount(self, schluessel):
        """Liest einen Betrag aus der Schaltfläche zurück."""

        text = self.amount_buttons[schluessel].text

        try:
            return float(
                text.replace(" €", "").replace(".", "").replace(",", ".")
            )
        except ValueError:
            return 0.0

    def _current_entry(self):

        return {
            "opening_balance": self._amount("opening_balance"),
            "income": self._amount("income"),
            "expenses": self._amount("expenses"),
            "closing_balance": self._amount("closing_balance"),
        }

    def _refresh_check(self):
        """Meldet Abweichungen schon beim Erfassen.

        Geprüft wird beides: die Rechnung innerhalb der Zeile und der
        Anschluss an den Eintrag davor. Beides erst in der Tabelle
        auffallen zu lassen, hieße, den Fehler zweimal suchen zu
        müssen.
        """

        eintrag = self._current_entry()

        meldungen = []

        if not self.db.cash_book_entry_is_valid(eintrag):

            erwartet = (
                eintrag["opening_balance"]
                + eintrag["income"]
                - eintrag["expenses"]
            )

            meldungen.append(
                f"Endbestand: erwartet {self.money(erwartet)}, "
                f"eingetragen {self.money(eintrag['closing_balance'])}"
            )

        vorheriger = self.db.get_previous_closing_balance(self.date_value)

        if vorheriger is not None and abs(
            eintrag["opening_balance"] - vorheriger
        ) > self.db.CASH_BOOK_TOLERANCE:

            meldungen.append(
                f"Startbestand: davor endete die Kasse mit "
                f"{self.money(vorheriger)}"
            )

        if not meldungen:
            self.check_label.text = "Rechnung geht auf."
            self.check_label.color = theme.TEXT_SECONDARY
            self.check_label.height = dp(34)
            return

        self.check_label.text = "Prüfen - " + "; ".join(meldungen)
        self.check_label.color = theme.ERROR

        # Zwei Befunde brauchen zwei Zeilen.
        self.check_label.height = dp(34 if len(meldungen) == 1 else 56)

    # =====================================================
    # Excel-Export
    # =====================================================

    def export_excel(self):
        """Schreibt den angezeigten Monat als Excel-Datei.

        Gedacht zum Ausdrucken und Abheften: Die letzte Spalte nennt
        bei auffälligen Zeilen den Grund im Klartext - auf Papier hilft
        ein rotes "Prüfen" ohne Erklärung niemandem weiter.
        """

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        eintraege, befunde = self.db.get_cash_book_entries_checked(
            self.selected_year, self.selected_month
        )

        if not eintraege:
            self.export_status.text = "Für diesen Monat gibt es nichts zu exportieren."
            return

        monat = MONTH_NAMES[self.selected_month - 1]

        workbook = Workbook()
        blatt = workbook.active
        blatt.title = f"{monat} {self.selected_year}"

        fett = Font(bold=True)

        blatt.append((f"Kassenbuch {monat} {self.selected_year}",))
        blatt["A1"].font = Font(bold=True, size=14)

        blatt.append(())

        ueberschriften = (
            "Datum", "Startbestand", "Einnahmen", "Ausgaben",
            "Endbestand", "Kommentar", "Prüfer", "Hinweis",
        )

        blatt.append(ueberschriften)

        for zelle in blatt[3]:
            zelle.font = fett

        auffaellig = 0

        for eintrag in eintraege:

            meldungen = befunde.get(eintrag["id"]) or []

            if meldungen:
                auffaellig += 1

            blatt.append((
                self.format_date(eintrag["entry_date"]),
                round(eintrag["opening_balance"] or 0, 2),
                round(eintrag["income"] or 0, 2),
                round(eintrag["expenses"] or 0, 2),
                round(eintrag["closing_balance"] or 0, 2),
                eintrag["comment"] or "",
                eintrag["auditor"] or "",
                "Prüfen: " + "; ".join(meldungen) if meldungen else "",
            ))

        summen = self.db.get_cash_book_totals(
            self.selected_year, self.selected_month
        )

        blatt.append(())

        summenzeile = blatt.max_row + 1

        blatt.append((
            "Summe",
            "",
            round(summen["income"], 2),
            round(summen["expenses"], 2),
            round(summen["closing_balance"], 2),
            "",
            "",
            (
                f"{auffaellig} {'Zeile' if auffaellig == 1 else 'Zeilen'} "
                "zu prüfen" if auffaellig else "alle Zeilen gehen auf"
            ),
        ))

        for zelle in blatt[summenzeile]:
            zelle.font = fett

        # Beträge als Währung, damit die Ausgabe ohne Nacharbeit
        # ausgedruckt werden kann.
        for zeile in blatt.iter_rows(min_row=4, min_col=2, max_col=5):
            for zelle in zeile:
                zelle.number_format = '#,##0.00 "€"'

        for spalte, breite in zip("ABCDEFGH", (12, 14, 12, 12, 14, 24, 14, 46)):
            blatt.column_dimensions[spalte].width = breite

        for zeile in blatt.iter_rows(min_row=4, min_col=8, max_col=8):
            for zelle in zeile:
                zelle.alignment = Alignment(wrap_text=True, vertical="top")

        # Fürs Drucken: quer, auf eine Seitenbreite, Kopfzeile
        # wiederholen.
        blatt.page_setup.orientation = "landscape"
        blatt.page_setup.fitToWidth = 1
        blatt.sheet_properties.pageSetUpPr.fitToPage = True
        blatt.print_title_rows = "3:3"

        dateiname = (
            f"kassenbuch_{self.selected_year}-{self.selected_month:02d}.xlsx"
        )

        ziel = storage.export_dir("excel") / dateiname

        workbook.save(ziel)

        self.letzte_ausgabe = ziel

        self.export_status.text = export_hinweis(ziel)

    # =====================================================
    # Speichern und löschen
    # =====================================================

    def save_entry(self):

        eintrag = self._current_entry()

        if self.selected_entry_id is None:

            neue_id = self.db.add_cash_book_entry(
                entry_date=self.date_value,
                comment=self.comment_input.text.strip(),
                auditor=self.auditor_input.text.strip(),
                **eintrag,
            )

            self.selected_entry_id = neue_id

            # Die Zeile gibt es jetzt - ein weiteres "Speichern" ändert
            # sie, statt eine zweite anzulegen. Das muss auch oben
            # stehen, sonst verspricht die Überschrift etwas anderes,
            # als der Knopf tut.
            self.form_title.text = "Zeile bearbeiten"

        else:

            self.db.update_cash_book_entry(
                entry_id=self.selected_entry_id,
                entry_date=self.date_value,
                comment=self.comment_input.text.strip(),
                auditor=self.auditor_input.text.strip(),
                **eintrag,
            )

        # Zum gespeicherten Datum springen, sonst verschwindet die
        # Zeile scheinbar spurlos, wenn sie in einen anderen Monat
        # gehört.
        gespeichert = date.fromisoformat(self.date_value)
        self.selected_year = gespeichert.year
        self.selected_month = gespeichert.month

        self._highlight_selection()
        self.refresh()

    def delete_entry(self):

        if self.selected_entry_id is None:
            return

        ConfirmPopup(
            title="Kassenbuch",
            message=(
                f"Zeile vom {self.format_date(self.date_value)} "
                "wirklich löschen?"
            ),
            on_confirm=self._delete_entry_confirmed,
        ).open()

    def _delete_entry_confirmed(self):

        self.db.delete_cash_book_entry(self.selected_entry_id)

        self.selected_entry_id = None
        self.new_entry()
        self.refresh()
