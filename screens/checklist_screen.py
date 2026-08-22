"""
=========================================================
KiG POS
=========================================================

Datei:
    screens/checklist_screen.py

Beschreibung:
    Checklisten für alles, was vor und nach einer
    Veranstaltung zu erledigen ist.

    Aufbau (Querformat):

        links    die angelegten Checklisten
        rechts   die Punkte der gewählten Liste

    Ein Punkt besteht aus Haken, Aufgabe und den
    Zusatzangaben Frist, Verantwortlich, Ansprechpartner
    und Infos (siehe widgets/checklists/checklist_item_row.py).

Version:
    1.0.0
=========================================================
"""

from datetime import datetime

from kivy.metrics import dp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.screenmanager import Screen

import config
import storage
import theme

from widgets.common.exporthinweis import (
    export_hinweis, hinweisfeld_vorbereiten,
)

from database import DatabaseManager
from widgets.checklists.checklist_item_row import ChecklistItemRow
from widgets.common.confirm_popup import ConfirmPopup
from widgets.common.date_picker_popup import DatePickerPopup
from widgets.common.kig_popup import KiGPopup
from widgets.common.rounded_input import RoundedInput
from widgets.common.rounded_panel import RoundedPanel
from widgets.kig_label import KiGLabel


class ChecklistScreen(Screen):

    LIST_WIDTH = 260
    LIST_BUTTON_HEIGHT = 54

    def __init__(self, **kwargs):

        super().__init__(**kwargs)

        self.db = DatabaseManager()

        self.selected_checklist_id = None
        self.checklist_buttons = {}

        # Im Hochformat stehen die Listen über den Punkten statt
        # daneben.
        self.hochformat = theme.is_portrait()

        root = BoxLayout(
            orientation="vertical" if self.hochformat else "horizontal",
            padding=dp(theme.SCREEN_PADDING),
            spacing=dp(theme.SCREEN_SPACING),
        )

        root.add_widget(self._build_list_panel())
        root.add_widget(self._build_items_panel())

        self.add_widget(root)

    # =====================================================
    # Links: die Checklisten
    # =====================================================

    def _build_list_panel(self):

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
            size_hint=(1, 0.3) if self.hochformat else (None, 1),
        )

        if not self.hochformat:
            panel.width = dp(self.LIST_WIDTH)

        panel.add_widget(self._title("Checklisten"))

        self.list_box = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.SPACE_XS),
            size_hint_y=None,
        )
        self.list_box.bind(minimum_height=self.list_box.setter("height"))

        scroll = ScrollView(do_scroll_x=False, bar_width=dp(8))
        scroll.add_widget(self.list_box)
        panel.add_widget(scroll)

        aktionen = BoxLayout(
            size_hint_y=None, height=dp(52), spacing=dp(theme.ROW_SPACING)
        )
        aktionen.add_widget(self._button("Neue Liste", self.new_checklist))
        aktionen.add_widget(self._button("Löschen", self.delete_checklist))
        panel.add_widget(aktionen)

        return panel

    def _refresh_checklists(self):

        self.list_box.clear_widgets()
        self.checklist_buttons = {}

        listen = self.db.get_checklists()

        if not listen:

            self.list_box.add_widget(Label(
                text="Noch keine Checkliste angelegt.",
                color=theme.TEXT_SECONDARY, font_size="14sp",
                size_hint_y=None, height=dp(46),
                halign="left", valign="middle", text_size=(None, dp(46)),
            ))

            self.selected_checklist_id = None

            return

        # Ist die gewählte Liste weg (gelöscht), rückt die erste nach.
        vorhandene = {liste["id"] for liste in listen}

        if self.selected_checklist_id not in vorhandene:
            self.selected_checklist_id = listen[0]["id"]

        for liste in listen:

            erledigt, gesamt = self.db.get_checklist_progress(liste["id"])

            beschriftung = liste["name"]

            if gesamt:
                beschriftung += f"   ({erledigt}/{gesamt})"

            button = Button(
                text=beschriftung, size_hint_y=None,
                height=dp(self.LIST_BUTTON_HEIGHT),
                background_normal="", background_down="",
                font_size="16sp", bold=True,
                halign="center", valign="middle",
            )

            button.bind(
                on_release=lambda _b, liste_id=liste["id"]:
                self.select_checklist(liste_id)
            )

            self.checklist_buttons[liste["id"]] = button
            self.list_box.add_widget(button)

        self._highlight_checklist()

    def _highlight_checklist(self):

        for liste_id, button in self.checklist_buttons.items():

            gewaehlt = liste_id == self.selected_checklist_id

            button.background_color = (
                theme.PRIMARY_ORANGE if gewaehlt else theme.SURFACE
            )
            button.color = theme.TEXT_WHITE if gewaehlt else theme.TEXT_PRIMARY

    def select_checklist(self, checklist_id):

        self.selected_checklist_id = checklist_id

        self._highlight_checklist()
        self._refresh_items()

    # =====================================================
    # Rechts: die Punkte
    # =====================================================

    def _build_items_panel(self):

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
            size_hint=(1, 0.7) if self.hochformat else (1, 1),
        )

        kopf = BoxLayout(size_hint_y=None, height=dp(40),
                         spacing=dp(theme.ROW_SPACING))

        self.items_title = self._title("Keine Liste gewählt")
        self.items_title.size_hint_y = 1
        kopf.add_widget(self.items_title)

        export_knopf = self._button("Excel exportieren", self.export_excel)
        export_knopf.size_hint_x = None
        export_knopf.width = dp(190)
        kopf.add_widget(export_knopf)

        panel.add_widget(kopf)

        self.status_label = Label(
            text="", color=theme.TEXT_SECONDARY, font_size="13sp",
            halign="left", valign="middle",
        )

        # Waechst mit: Nach einem Export steht hier zusaetzlich der
        # Ordner (siehe widgets/common/exporthinweis.py).
        hinweisfeld_vorbereiten(self.status_label, dp(24))
        panel.add_widget(self.status_label)

        # Kopfzeile über den Spalten - ohne sie ist nicht zu erraten,
        # was in welches Feld gehört.
        self.items_header = BoxLayout(
            size_hint_y=None, height=dp(26),
            spacing=dp(theme.ROW_SPACING),
        )

        for text, breite, feste_breite in (
            ("", ChecklistItemRow.HAKEN_WIDTH, True),
            ("Aufgabe", 0.34, False),
            ("Frist", 0.14, False),
            ("Verantwortlich", 0.16, False),
            ("Ansprechpartner", 0.16, False),
            ("Infos", 0.20, False),
            ("", ChecklistItemRow.REMOVE_WIDTH, True),
        ):
            label = Label(
                text=text, bold=True, color=theme.TEXT_SECONDARY,
                font_size="13sp", halign="left", valign="middle",
                text_size=(None, dp(26)),
            )

            if feste_breite:
                label.size_hint_x = None
                label.width = dp(breite)
            else:
                label.size_hint_x = breite

            self.items_header.add_widget(label)

        panel.add_widget(self.items_header)

        self.items_box = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.ROW_SPACING),
            size_hint_y=None,
        )
        self.items_box.bind(minimum_height=self.items_box.setter("height"))

        scroll = ScrollView(do_scroll_x=False, bar_width=dp(10))
        scroll.add_widget(self.items_box)
        panel.add_widget(scroll)

        # Neue Aufgabe anlegen
        neue_zeile = BoxLayout(
            size_hint_y=None, height=dp(56), spacing=dp(theme.ROW_SPACING)
        )

        self.new_task_input = RoundedInput(
            hint_text="Neue Aufgabe eintragen", multiline=False,
        )
        self.new_task_input.foreground_color = theme.INPUT_TEXT
        self.new_task_input.hint_text_color = theme.INPUT_HINT
        self.new_task_input.bind(on_text_validate=lambda *_a: self.add_item())

        neue_zeile.add_widget(self.new_task_input)

        hinzufuegen = self._button(
            "Hinzufügen", self.add_item,
            background=theme.PRIMARY_ORANGE, color=theme.TEXT_WHITE,
        )
        hinzufuegen.size_hint_x = None
        hinzufuegen.width = dp(150)
        neue_zeile.add_widget(hinzufuegen)

        panel.add_widget(neue_zeile)

        return panel

    def _refresh_items(self):

        self.items_box.clear_widgets()

        if self.selected_checklist_id is None:

            self.items_title.text = "Keine Liste gewählt"
            self.status_label.text = (
                "Links über \"Neue Liste\" eine Checkliste anlegen."
            )
            self.items_header.opacity = 0

            return

        self.items_header.opacity = 1

        liste = self.db.get_checklist(self.selected_checklist_id)

        self.items_title.text = liste["name"] if liste else "Checkliste"

        punkte = self.db.get_checklist_items(self.selected_checklist_id)

        if not punkte:

            self.items_box.add_widget(Label(
                text="Noch keine Aufgabe eingetragen.",
                color=theme.TEXT_SECONDARY, font_size="14sp",
                size_hint_y=None, height=dp(46),
                halign="left", valign="middle", text_size=(None, dp(46)),
            ))

        else:

            for punkt in punkte:

                self.items_box.add_widget(ChecklistItemRow(
                    item=punkt,
                    on_toggle=self.item_toggled,
                    on_change=self.item_changed,
                    on_deadline=self.open_deadline_picker,
                    on_remove=self.remove_item,
                ))

        self._refresh_status()

    def _refresh_status(self):

        if self.selected_checklist_id is None:
            return

        erledigt, gesamt = self.db.get_checklist_progress(
            self.selected_checklist_id
        )

        if not gesamt:
            self.status_label.text = "Noch nichts einzutragen."
            self.status_label.color = theme.TEXT_SECONDARY
            return

        offen = gesamt - erledigt

        if offen:
            self.status_label.text = (
                f"{erledigt} von {gesamt} erledigt   ·   "
                f"{offen} {'Aufgabe' if offen == 1 else 'Aufgaben'} offen"
            )
            self.status_label.color = theme.TEXT_SECONDARY
        else:
            self.status_label.text = f"Alle {gesamt} Aufgaben erledigt."
            self.status_label.color = theme.SUCCESS

    # =====================================================
    # Punkte ändern
    # =====================================================

    def item_toggled(self, row, done):

        self.db.update_checklist_item(row.item["id"], done=done)

        self._refresh_status()

        # Der Fortschritt steht auch an der Liste links.
        self._refresh_checklist_labels()

    def item_changed(self, row, feld, wert):

        gespeichert = row.item[feld] or ""

        if wert == gespeichert:
            return

        self.db.update_checklist_item(row.item["id"], **{feld: wert})

        row.item = dict(row.item)
        row.item[feld] = wert

    def open_deadline_picker(self, row):

        DatePickerPopup(
            title="Frist",
            initial_date=row.item["deadline"],
            on_select=row.set_deadline,
        ).open()

    def remove_item(self, row):

        ConfirmPopup(
            title="Checkliste",
            message=f"\"{row.item['task']}\" aus der Liste entfernen?",
            on_confirm=lambda: self._remove_item_confirmed(row),
        ).open()

    def _remove_item_confirmed(self, row):

        self.db.delete_checklist_item(row.item["id"])

        self._refresh_items()
        self._refresh_checklist_labels()

    def add_item(self):

        if self.selected_checklist_id is None:
            return

        aufgabe = self.new_task_input.text.strip()

        if not aufgabe:
            return

        self.db.add_checklist_item(self.selected_checklist_id, aufgabe)

        self.new_task_input.text = ""

        self._refresh_items()
        self._refresh_checklist_labels()

    def _refresh_checklist_labels(self):
        """Aktualisiert nur die Beschriftungen links (Fortschritt),
        ohne die Auswahl zu verlieren."""

        gewaehlt = self.selected_checklist_id

        self._refresh_checklists()

        self.selected_checklist_id = gewaehlt
        self._highlight_checklist()

    # =====================================================
    # Checklisten anlegen und löschen
    # =====================================================

    def new_checklist(self):

        inhalt = BoxLayout(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
        )

        feld = RoundedInput(
            hint_text="z. B. Stadtfest 2026", multiline=False,
            size_hint_y=None, height=dp(56),
        )
        feld.foreground_color = theme.INPUT_TEXT
        feld.hint_text_color = theme.INPUT_HINT

        inhalt.add_widget(Label(
            text="Name der neuen Checkliste", color=theme.TEXT_PRIMARY,
            font_size="16sp", size_hint_y=None, height=dp(34),
        ))
        inhalt.add_widget(feld)

        popup = KiGPopup(
            title="Neue Checkliste", content=inhalt,
            size_hint=(0.55, None), height=dp(280), auto_dismiss=False,
        )

        knoepfe = BoxLayout(
            size_hint_y=None, height=dp(52), spacing=dp(theme.ROW_SPACING)
        )
        knoepfe.add_widget(self._button("Abbrechen", popup.dismiss))
        knoepfe.add_widget(self._button(
            "Anlegen",
            lambda: (popup.dismiss(), self._create_checklist(feld.text)),
            background=theme.PRIMARY_ORANGE, color=theme.TEXT_WHITE,
        ))
        inhalt.add_widget(knoepfe)

        popup.open()

        feld.focus = True

    def _create_checklist(self, name):

        neue_id = self.db.add_checklist(name)

        if neue_id is None:
            return

        self.selected_checklist_id = neue_id

        self._refresh_checklists()
        self._refresh_items()

    def delete_checklist(self):

        if self.selected_checklist_id is None:
            return

        liste = self.db.get_checklist(self.selected_checklist_id)

        if liste is None:
            return

        _erledigt, gesamt = self.db.get_checklist_progress(liste["id"])

        ConfirmPopup(
            title="Checkliste löschen",
            message=(
                f"\"{liste['name']}\" mit {gesamt} "
                f"{'Aufgabe' if gesamt == 1 else 'Aufgaben'} löschen?"
            ),
            on_confirm=self._delete_checklist_confirmed,
        ).open()

    def _delete_checklist_confirmed(self):

        self.db.delete_checklist(self.selected_checklist_id)

        self.selected_checklist_id = None

        self._refresh_checklists()
        self._refresh_items()

    # =====================================================
    # Excel-Export
    # =====================================================

    def export_excel(self):
        """Schreibt die gewählte Checkliste als Excel-Datei.

        Zum Ausdrucken und Mitnehmen: Wer am Stand steht, hat selten
        das Tablet in der Hand.
        """

        if self.selected_checklist_id is None:
            self.status_label.text = "Keine Checkliste gewählt."
            return

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        liste = self.db.get_checklist(self.selected_checklist_id)
        punkte = self.db.get_checklist_items(self.selected_checklist_id)

        if not punkte:
            self.status_label.text = "Diese Liste ist noch leer."
            return

        workbook = Workbook()
        blatt = workbook.active
        blatt.title = liste["name"][:31] or "Checkliste"

        fett = Font(bold=True)

        blatt.append((liste["name"],))
        blatt["A1"].font = Font(bold=True, size=14)

        blatt.append(())

        blatt.append((
            "Erledigt", "Aufgabe", "Frist", "Verantwortlich",
            "Ansprechpartner", "Infos",
        ))

        for zelle in blatt[3]:
            zelle.font = fett

        for punkt in punkte:

            blatt.append((
                "x" if punkt["done"] else "",
                punkt["task"],
                ChecklistItemRow.format_deadline(punkt["deadline"])
                if punkt["deadline"] else "",
                punkt["responsible"] or "",
                punkt["contact"] or "",
                punkt["info"] or "",
            ))

        erledigt, gesamt = self.db.get_checklist_progress(liste["id"])

        blatt.append(())
        blatt.append((f"{erledigt} von {gesamt} erledigt",))
        blatt[blatt.max_row][0].font = fett

        for spalte, breite in zip("ABCDEF", (10, 42, 12, 20, 20, 34)):
            blatt.column_dimensions[spalte].width = breite

        for zeile in blatt.iter_rows(min_row=4, min_col=2, max_col=6):
            for zelle in zeile:
                zelle.alignment = Alignment(wrap_text=True, vertical="top")

        blatt.page_setup.orientation = "landscape"
        blatt.page_setup.fitToWidth = 1
        blatt.sheet_properties.pageSetUpPr.fitToPage = True
        blatt.print_title_rows = "3:3"

        sicherer_name = "".join(
            zeichen if zeichen.isalnum() or zeichen in " -_" else "_"
            for zeichen in liste["name"]
        ).strip().replace(" ", "_")

        dateiname = (
            f"checkliste_{sicherer_name or 'liste'}_"
            f"{datetime.now():%Y-%m-%d}.xlsx"
        )

        ziel = storage.export_dir("excel") / dateiname

        workbook.save(ziel)

        self.status_label.text = export_hinweis(ziel)
        self.status_label.color = theme.TEXT_SECONDARY

    # =====================================================
    # Screen wird geöffnet
    # =====================================================

    def on_pre_enter(self, *_args):

        self._refresh_checklists()
        self._refresh_items()

    # =====================================================
    # Hilfsmittel
    # =====================================================

    @staticmethod
    def _button(text, callback, background=None, color=None):

        button = Button(
            text=text, background_normal="", background_down="",
            background_color=background or theme.SURFACE,
            color=color or theme.TEXT_PRIMARY,
            font_size="15sp", bold=True,
        )
        button.bind(on_release=lambda *_args: callback())

        return button

    @staticmethod
    def _title(text):

        label = KiGLabel(text=text)
        label.set_font_size(24)
        label.set_bold(True)
        label.set_alignment("left")
        label.set_color(theme.PRIMARY_ORANGE)
        label.size_hint_y = None
        label.height = dp(36)

        return label
