"""
=========================================================
KiG POS
=========================================================

Datei:
    screens/shift_plan_screen.py

Beschreibung:
    Schichtplan einer Veranstaltung.

    Beantwortet eine Frage: Wo fehlen noch Helfer?

    Aufbau (Querformat):

        links    die Veranstaltungen mit Schichtplan
        rechts   die Schichten der gewählten Veranstaltung

    Je Schicht steht dort Tätigkeit, Zeit, wie viele
    gebraucht werden und wie viele eingetragen sind. Der
    Balken daneben zeigt dasselbe auf einen Blick:

        grün      besetzt
        orange    teilweise besetzt
        rot       noch niemand

    Die Zahl "eingetragen" wird nirgends gepflegt - sie ist
    die Anzahl der Helfer in der Schicht (siehe
    database.py:_create_shift_tables).

Version:
    1.0.0
=========================================================
"""

from datetime import datetime

from kivy.metrics import dp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.screenmanager import Screen

import config
import storage
import teilen
import theme

from database import DatabaseManager
from widgets.common.confirm_popup import ConfirmPopup
from widgets.common.exporthinweis import (
    export_hinweis, hinweisfeld_vorbereiten,
)
from widgets.common.kig_popup import KiGPopup
from widgets.common.kig_symbol import KREUZ, KiGSymbolButton
from widgets.common.numpad.numpad_popup import NumpadPopup
from widgets.common.rounded_input import RoundedInput
from widgets.common.rounded_panel import RoundedPanel
from widgets.kig_label import KiGLabel
from widgets.shifts.shift_row import ShiftRow


class ShiftPlanScreen(Screen):

    LIST_WIDTH = 260
    LIST_BUTTON_HEIGHT = 60

    def __init__(self, **kwargs):

        super().__init__(**kwargs)

        self.db = DatabaseManager()

        self.selected_plan_id = None
        self.plan_buttons = {}
        self.shift_rows = {}

        # Zuletzt ausgegebene Datei - sie haengt am Teilen-Knopf.
        self.letzte_ausgabe = None

        # Im Hochformat stehen die Veranstaltungen über den Schichten
        # statt daneben.
        self.hochformat = theme.is_portrait()

        root = BoxLayout(
            orientation="vertical" if self.hochformat else "horizontal",
            padding=dp(theme.SCREEN_PADDING),
            spacing=dp(theme.SCREEN_SPACING),
        )

        root.add_widget(self._build_list_panel())
        root.add_widget(self._build_shifts_panel())

        self.add_widget(root)

    # =====================================================
    # Links: die Veranstaltungen
    # =====================================================

    def _build_list_panel(self):

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
            size_hint=(1, 0.3) if self.hochformat else (None, 1),
        )

        if not self.hochformat:
            panel.width = dp(self.LIST_WIDTH)

        panel.add_widget(self._title("Schichtpläne"))

        self.list_box = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.SPACE_XS),
            size_hint_y=None,
        )
        self.list_box.bind(minimum_height=self.list_box.setter("height"))

        scroll = ScrollView(do_scroll_x=False, bar_width=dp(8))
        scroll.add_widget(self.list_box)
        panel.add_widget(scroll)

        aktionen = BoxLayout(
            size_hint_y=None, height=dp(52), spacing=dp(theme.ROW_SPACING)
        )
        aktionen.add_widget(self._button("Plan anlegen", self.new_plan))
        aktionen.add_widget(self._button("Löschen", self.delete_plan))
        panel.add_widget(aktionen)

        return panel

    def _refresh_plans(self):

        self.list_box.clear_widgets()
        self.plan_buttons = {}

        plaene = self.db.get_shift_plans()

        if not plaene:

            hinweis = Label(
                text=(
                    "Noch kein Schichtplan.\n"
                    "Beim Anlegen einer Veranstaltung im Kalender "
                    "ankreuzen - oder hier über \"Plan anlegen\"."
                ),
                color=theme.TEXT_SECONDARY, font_size="14sp",
                size_hint_y=None, height=dp(90),
                halign="left", valign="top",
            )

            # Umbruch an der Panelbreite statt an einer festen Zahl -
            # auf einem Telefon lief der Text sonst links und rechts
            # aus der Karte heraus.
            hinweis.bind(
                size=lambda instanz, groesse: setattr(
                    instanz, "text_size", groesse
                )
            )

            self.list_box.add_widget(hinweis)

            self.selected_plan_id = None

            return

        vorhandene = {plan["id"] for plan in plaene}

        if self.selected_plan_id not in vorhandene:
            self.selected_plan_id = plaene[0]["id"]

        for plan in plaene:

            besetzt, plaetze, offen = self.db.get_shift_plan_summary(
                plan["id"]
            )

            beschriftung = f"{plan['event_name']}\n{self.format_date(plan['event_date'])}"

            if plaetze:
                beschriftung += f"   ({besetzt}/{plaetze})"

            button = Button(
                text=beschriftung, size_hint_y=None,
                height=dp(self.LIST_BUTTON_HEIGHT),
                background_normal="", background_down="",
                font_size="15sp", bold=True,
                halign="center", valign="middle",
            )
            button.bind(
                size=lambda instanz, groesse: setattr(
                    instanz, "text_size", groesse
                )
            )

            button.bind(
                on_release=lambda _b, plan_id=plan["id"]:
                self.select_plan(plan_id)
            )

            self.plan_buttons[plan["id"]] = button
            self.list_box.add_widget(button)

        self._highlight_plan()

    def _highlight_plan(self):

        for plan_id, button in self.plan_buttons.items():

            gewaehlt = plan_id == self.selected_plan_id

            button.background_color = (
                theme.PRIMARY_ORANGE if gewaehlt else theme.SURFACE
            )
            button.color = theme.TEXT_WHITE if gewaehlt else theme.TEXT_PRIMARY

    def select_plan(self, plan_id):

        self.selected_plan_id = plan_id

        self._highlight_plan()
        self._refresh_shifts()

    # =====================================================
    # Rechts: die Schichten
    # =====================================================

    def _build_shifts_panel(self):

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
            size_hint=(1, 0.7) if self.hochformat else (1, 1),
        )

        self.shifts_title = self._title("Keine Veranstaltung gewählt")

        # Drei Beschriftungen nebeneinander passen auf ein Telefon nur
        # gekuerzt - ausgeschrieben ueberlappten sie sich.
        schmal = theme.is_narrow()

        uebernehmen = self._button(
            "Übernehmen" if schmal else "Schichten übernehmen",
            self.copy_shifts,
        )
        export_knopf = self._button(
            "Export" if schmal else "Excel exportieren", self.export_excel
        )
        teilen_knopf = self._button("Teilen", self.teilen_clicked)

        if self.hochformat:

            # Auf dem Telefon passen Überschrift und drei Knöpfe nicht
            # nebeneinander - dort stehen die Knöpfe darunter und
            # teilen sich die Breite.
            kopf = BoxLayout(
                orientation="vertical", size_hint_y=None,
                height=dp(36 + 52 + theme.ROW_SPACING),
                spacing=dp(theme.ROW_SPACING),
            )

            kopf.add_widget(self.shifts_title)

            knopfreihe = BoxLayout(
                size_hint_y=None, height=dp(52),
                spacing=dp(theme.ROW_SPACING),
            )
            knopfreihe.add_widget(uebernehmen)
            knopfreihe.add_widget(export_knopf)
            knopfreihe.add_widget(teilen_knopf)

            kopf.add_widget(knopfreihe)

        else:

            kopf = BoxLayout(size_hint_y=None, height=dp(40),
                             spacing=dp(theme.ROW_SPACING))

            self.shifts_title.size_hint_y = 1
            kopf.add_widget(self.shifts_title)

            uebernehmen.size_hint_x = None
            uebernehmen.width = dp(210)
            kopf.add_widget(uebernehmen)

            export_knopf.size_hint_x = None
            export_knopf.width = dp(190)
            kopf.add_widget(export_knopf)

            teilen_knopf.size_hint_x = None
            teilen_knopf.width = dp(110)
            kopf.add_widget(teilen_knopf)

        panel.add_widget(kopf)

        self.status_label = Label(
            text="", color=theme.TEXT_SECONDARY, font_size="14sp",
            halign="left", valign="middle", bold=True,
        )

        hinweisfeld_vorbereiten(self.status_label, dp(26))
        panel.add_widget(self.status_label)

        # Kopfzeile über den Spalten - dieselbe Aufteilung wie in
        # ShiftRow, nur gemeinsam ändern.
        self.shifts_header = BoxLayout(
            size_hint_y=None, height=dp(26),
            spacing=dp(theme.ROW_SPACING),
        )

        for text, breite, feste_breite in (
            ("Tätigkeit", ShiftRow.TASK_WIDTH, False),
            ("von", ShiftRow.TIME_WIDTH, False),
            ("bis", ShiftRow.TIME_WIDTH, False),
            ("Ist / Soll", ShiftRow.COUNT_WIDTH, True),
            ("Besetzung", ShiftRow.BAR_WIDTH, False),
            ("Helfer", ShiftRow.HELPER_WIDTH, False),
            ("", ShiftRow.REMOVE_WIDTH, True),
        ):
            label = Label(
                text=text, bold=True, color=theme.TEXT_SECONDARY,
                font_size="13sp", halign="left", valign="middle",
                text_size=(None, dp(26)),
            )

            if feste_breite:
                label.size_hint_x = None
                label.width = dp(breite)
            else:
                label.size_hint_x = breite

            self.shifts_header.add_widget(label)

        panel.add_widget(self.shifts_header)

        self.shifts_box = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.ROW_SPACING),
            size_hint_y=None,
        )
        self.shifts_box.bind(minimum_height=self.shifts_box.setter("height"))

        scroll = ScrollView(do_scroll_x=False, bar_width=dp(10))
        scroll.add_widget(self.shifts_box)
        panel.add_widget(scroll)

        # Neue Schicht anlegen
        neue_zeile = BoxLayout(
            size_hint_y=None, height=dp(56), spacing=dp(theme.ROW_SPACING)
        )

        self.new_task_input = RoundedInput(
            hint_text="Neue Schicht: Tätigkeit eintragen", multiline=False,
        )
        self.new_task_input.foreground_color = theme.INPUT_TEXT
        self.new_task_input.hint_text_color = theme.INPUT_HINT
        self.new_task_input.bind(on_text_validate=lambda *_a: self.add_shift())

        neue_zeile.add_widget(self.new_task_input)

        hinzufuegen = self._button(
            "Hinzufügen", self.add_shift,
            background=theme.PRIMARY_ORANGE, color=theme.TEXT_WHITE,
        )
        hinzufuegen.size_hint_x = None
        hinzufuegen.width = dp(150)
        neue_zeile.add_widget(hinzufuegen)

        panel.add_widget(neue_zeile)

        return panel

    def _refresh_shifts(self):

        self.shifts_box.clear_widgets()
        self.shift_rows = {}

        if self.selected_plan_id is None:

            self.shifts_title.text = "Keine Veranstaltung gewählt"
            self.status_label.text = (
                f"{'Oben' if self.hochformat else 'Links'} über "
                f"\"Plan anlegen\" einen Schichtplan anlegen."
            )
            self.status_label.color = theme.TEXT_SECONDARY
            self.shifts_header.opacity = 0

            return

        self.shifts_header.opacity = 1

        plan = self.db.get_shift_plan(self.selected_plan_id)

        self.shifts_title.text = (
            f"{plan['event_name']}   {self.format_date(plan['event_date'])}"
            if plan else "Schichtplan"
        )

        schichten = self.db.get_shifts(self.selected_plan_id)

        if not schichten:

            self.shifts_box.add_widget(Label(
                text="Noch keine Schicht eingetragen.",
                color=theme.TEXT_SECONDARY, font_size="14sp",
                size_hint_y=None, height=dp(46),
                halign="left", valign="middle", text_size=(None, dp(46)),
            ))

        else:

            for schicht in schichten:

                zeile = ShiftRow(
                    shift=self._mit_helfern(schicht),
                    on_change=self.shift_changed,
                    on_needed=self.open_needed_numpad,
                    on_helpers=self.open_helpers,
                    on_remove=self.remove_shift,
                )

                self.shift_rows[schicht["id"]] = zeile
                self.shifts_box.add_widget(zeile)

        self._refresh_status()

    def _mit_helfern(self, schicht):
        """Ergänzt die Schicht um die Namen ihrer Helfer."""

        daten = dict(schicht)

        namen = [
            helfer["name"]
            for helfer in self.db.get_shift_helpers(schicht["id"])
        ]

        daten["helfer_namen"] = ", ".join(namen)

        return daten

    def _refresh_status(self):

        if self.selected_plan_id is None:
            return

        besetzt, plaetze, offen = self.db.get_shift_plan_summary(
            self.selected_plan_id
        )

        if not plaetze:
            self.status_label.text = "Noch keine Schicht eingetragen."
            self.status_label.color = theme.TEXT_SECONDARY
            return

        text = f"{besetzt} von {plaetze} Plätzen besetzt"

        if offen:
            self.status_label.text = (
                f"{text}   ·   {offen} "
                f"{'Schicht braucht' if offen == 1 else 'Schichten brauchen'} "
                f"noch Helfer"
            )
            self.status_label.color = (
                theme.ERROR if besetzt == 0 else theme.WARNING
            )
        else:
            self.status_label.text = f"{text}   ·   alle Schichten besetzt"
            self.status_label.color = theme.SUCCESS

    def _refresh_plan_labels(self):
        """Aktualisiert die Beschriftungen links, ohne die Auswahl zu
        verlieren."""

        gewaehlt = self.selected_plan_id

        self._refresh_plans()

        self.selected_plan_id = gewaehlt
        self._highlight_plan()

    # =====================================================
    # Schichten ändern
    # =====================================================

    def shift_changed(self, row, feld, wert):

        gespeichert = row.shift[feld] or ""

        wert = (wert or "").strip()

        if wert == gespeichert:
            return

        self.db.update_shift(row.shift["id"], **{feld: wert})

        row.shift = dict(row.shift)
        row.shift[feld] = wert

    def open_needed_numpad(self, row):
        """Wie viele Helfer braucht diese Schicht?"""

        NumpadPopup(
            title=f"{row.shift['task'] or 'Schicht'}: benötigte Helfer",
            value=int(row.shift["needed"] or 0),
            mode="count",
            on_confirm=lambda anzahl: self.needed_entered(row, anzahl),
        ).open()

    def needed_entered(self, row, anzahl):

        self.db.update_shift(row.shift["id"], needed=anzahl)

        self._aktualisiere_zeile(row)

    def remove_shift(self, row):

        ConfirmPopup(
            title="Schicht entfernen",
            message=(
                f"\"{row.shift['task'] or 'Schicht'}\" mit "
                f"{row.shift['besetzt']} "
                f"{'Helfer' if row.shift['besetzt'] == 1 else 'Helfern'} "
                f"entfernen?"
            ),
            on_confirm=lambda: self._remove_shift_confirmed(row),
        ).open()

    def _remove_shift_confirmed(self, row):

        self.db.delete_shift(row.shift["id"])

        self._refresh_shifts()
        self._refresh_plan_labels()

    def add_shift(self):

        if self.selected_plan_id is None:
            return

        taetigkeit = self.new_task_input.text.strip()

        if not taetigkeit:
            return

        self.db.add_shift(self.selected_plan_id, taetigkeit)

        self.new_task_input.text = ""

        self._refresh_shifts()
        self._refresh_plan_labels()

    def _aktualisiere_zeile(self, row):
        """Holt eine einzelne Zeile frisch aus der Datenbank."""

        for schicht in self.db.get_shifts(self.selected_plan_id):

            if schicht["id"] == row.shift["id"]:
                row.aktualisieren(self._mit_helfern(schicht))
                break

        self._refresh_status()
        self._refresh_plan_labels()

    # =====================================================
    # Helfer
    # =====================================================

    def open_helpers(self, row):
        """Wer ist für diese Schicht eingetragen?"""

        inhalt = BoxLayout(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
        )

        liste = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.SPACE_XS),
            size_hint_y=None,
        )
        liste.bind(minimum_height=liste.setter("height"))

        scroll = ScrollView(do_scroll_x=False, bar_width=dp(8))
        scroll.add_widget(liste)
        inhalt.add_widget(scroll)

        feld = RoundedInput(
            hint_text="Name eintragen", multiline=False,
            size_hint_y=None, height=dp(56),
        )
        feld.foreground_color = theme.INPUT_TEXT
        feld.hint_text_color = theme.INPUT_HINT

        popup = KiGPopup(
            title=f"Helfer: {row.shift['task'] or 'Schicht'}",
            content=inhalt,
            size_hint=(0.55, None), height=dp(520), auto_dismiss=False,
        )

        def neu_zeichnen():

            liste.clear_widgets()

            helfer = self.db.get_shift_helpers(row.shift["id"])

            if not helfer:

                liste.add_widget(Label(
                    text="Noch niemand eingetragen.",
                    color=theme.TEXT_SECONDARY, font_size="14sp",
                    size_hint_y=None, height=dp(46),
                    halign="left", valign="middle",
                    text_size=(None, dp(46)),
                ))

            for eintrag in helfer:

                zeile = BoxLayout(
                    size_hint_y=None, height=dp(52),
                    spacing=dp(theme.ROW_SPACING),
                )

                name = Label(
                    text=eintrag["name"], color=theme.TEXT_PRIMARY,
                    font_size="16sp", halign="left", valign="middle",
                )
                name.bind(
                    size=lambda instanz, groesse: setattr(
                        instanz, "text_size", groesse
                    )
                )
                zeile.add_widget(name)

                entfernen = KiGSymbolButton(
                    symbol=KREUZ, size_hint_x=None, width=dp(52),
                    symbol_color=theme.ERROR,
                )
                entfernen.bind(
                    on_release=lambda _b, helfer_id=eintrag["id"]:
                    entfernen_bestaetigt(helfer_id)
                )
                zeile.add_widget(entfernen)

                liste.add_widget(zeile)

        def entfernen_bestaetigt(helfer_id):

            self.db.delete_shift_helper(helfer_id)

            neu_zeichnen()
            self._aktualisiere_zeile(row)

        def eintragen():

            if not feld.text.strip():
                return

            self.db.add_shift_helper(row.shift["id"], feld.text)

            feld.text = ""

            neu_zeichnen()
            self._aktualisiere_zeile(row)

        feld.bind(on_text_validate=lambda *_a: eintragen())

        eingabe = BoxLayout(
            size_hint_y=None, height=dp(56), spacing=dp(theme.ROW_SPACING)
        )
        eingabe.add_widget(feld)

        hinzufuegen = self._button(
            "Eintragen", eintragen,
            background=theme.PRIMARY_ORANGE, color=theme.TEXT_WHITE,
        )
        hinzufuegen.size_hint_x = None
        hinzufuegen.width = dp(150)
        eingabe.add_widget(hinzufuegen)

        inhalt.add_widget(eingabe)

        schliessen = self._button("Schließen", popup.dismiss)
        schliessen.size_hint_y = None
        schliessen.height = dp(52)
        inhalt.add_widget(schliessen)

        neu_zeichnen()

        popup.open()

        feld.focus = True

    # =====================================================
    # Pläne anlegen und löschen
    # =====================================================

    def new_plan(self):
        """Legt einen Plan für eine Veranstaltung an, die noch keinen
        hat."""

        offene = self.db.get_events_without_shift_plan()

        inhalt = BoxLayout(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
        )

        popup = KiGPopup(
            title="Schichtplan anlegen", content=inhalt,
            size_hint=(0.55, None), height=dp(480), auto_dismiss=False,
        )

        if not offene:

            inhalt.add_widget(Label(
                text=(
                    "Alle Veranstaltungen haben bereits einen Plan.\n\n"
                    "Neue Veranstaltungen werden im Kalender angelegt."
                ),
                color=theme.TEXT_SECONDARY, font_size="15sp",
                halign="left", valign="top",
            ))

        else:

            inhalt.add_widget(Label(
                text="Für welche Veranstaltung?",
                color=theme.TEXT_PRIMARY, font_size="16sp",
                size_hint_y=None, height=dp(34),
            ))

            liste = BoxLayout(
                orientation="vertical",
                spacing=dp(theme.SPACE_XS),
                size_hint_y=None,
            )
            liste.bind(minimum_height=liste.setter("height"))

            scroll = ScrollView(do_scroll_x=False, bar_width=dp(8))
            scroll.add_widget(liste)
            inhalt.add_widget(scroll)

            for veranstaltung in offene:

                knopf = self._button(
                    f"{veranstaltung['name']}   "
                    f"{self.format_date(veranstaltung['start_date'])}",
                    lambda event_id=veranstaltung["id"]: (
                        popup.dismiss(), self._create_plan(event_id)
                    ),
                )
                knopf.size_hint_y = None
                knopf.height = dp(52)

                liste.add_widget(knopf)

        schliessen = self._button("Schließen", popup.dismiss)
        schliessen.size_hint_y = None
        schliessen.height = dp(52)
        inhalt.add_widget(schliessen)

        popup.open()

    def _create_plan(self, event_id):

        neue_id = self.db.add_shift_plan(event_id)

        if neue_id is None:
            return

        self.selected_plan_id = neue_id

        self._refresh_plans()
        self._refresh_shifts()

    def delete_plan(self):

        if self.selected_plan_id is None:
            return

        plan = self.db.get_shift_plan(self.selected_plan_id)

        if plan is None:
            return

        _besetzt, plaetze, _offen = self.db.get_shift_plan_summary(plan["id"])

        ConfirmPopup(
            title="Schichtplan löschen",
            message=(
                f"Schichtplan für \"{plan['event_name']}\" mit {plaetze} "
                f"{'Platz' if plaetze == 1 else 'Plätzen'} löschen?\n\n"
                f"Die Veranstaltung selbst bleibt im Kalender."
            ),
            on_confirm=self._delete_plan_confirmed,
        ).open()

    def _delete_plan_confirmed(self):

        self.db.delete_shift_plan(self.selected_plan_id)

        self.selected_plan_id = None

        self._refresh_plans()
        self._refresh_shifts()

    def copy_shifts(self):
        """Übernimmt die Schichten einer anderen Veranstaltung.

        Die Tätigkeiten sind jedes Jahr dieselben - nur die Helfer
        sind andere. Übernommen wird deshalb das Gerüst, nicht wer
        letztes Mal da war.
        """

        if self.selected_plan_id is None:
            return

        andere = [
            plan for plan in self.db.get_shift_plans()
            if plan["id"] != self.selected_plan_id
        ]

        inhalt = BoxLayout(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
        )

        popup = KiGPopup(
            title="Schichten übernehmen", content=inhalt,
            size_hint=(0.55, None), height=dp(480), auto_dismiss=False,
        )

        if not andere:

            inhalt.add_widget(Label(
                text=(
                    "Es gibt noch keinen zweiten Schichtplan, aus dem "
                    "sich etwas übernehmen ließe."
                ),
                color=theme.TEXT_SECONDARY, font_size="15sp",
                halign="left", valign="top",
            ))

        else:

            inhalt.add_widget(Label(
                text="Schichten übernehmen aus:",
                color=theme.TEXT_PRIMARY, font_size="16sp",
                size_hint_y=None, height=dp(34),
            ))

            liste = BoxLayout(
                orientation="vertical",
                spacing=dp(theme.SPACE_XS),
                size_hint_y=None,
            )
            liste.bind(minimum_height=liste.setter("height"))

            scroll = ScrollView(do_scroll_x=False, bar_width=dp(8))
            scroll.add_widget(liste)
            inhalt.add_widget(scroll)

            for plan in andere:

                knopf = self._button(
                    f"{plan['event_name']}   "
                    f"{self.format_date(plan['event_date'])}",
                    lambda plan_id=plan["id"]: (
                        popup.dismiss(), self._copy_shifts_from(plan_id)
                    ),
                )
                knopf.size_hint_y = None
                knopf.height = dp(52)

                liste.add_widget(knopf)

        schliessen = self._button("Schließen", popup.dismiss)
        schliessen.size_hint_y = None
        schliessen.height = dp(52)
        inhalt.add_widget(schliessen)

        popup.open()

    def _copy_shifts_from(self, quell_plan_id):

        uebernommen = self.db.copy_shifts(
            quell_plan_id, self.selected_plan_id
        )

        self._refresh_shifts()
        self._refresh_plan_labels()

        if uebernommen:
            self.status_label.text = (
                f"{uebernommen} "
                f"{'Schicht' if uebernommen == 1 else 'Schichten'} "
                f"übernommen - Helfer bitte neu eintragen."
            )
            self.status_label.color = theme.TEXT_SECONDARY

    # =====================================================
    # Excel-Export
    # =====================================================

    def export_excel(self):
        """Schreibt den Schichtplan als Excel-Datei.

        Zum Aushängen: Am Stand hat selten jemand das Tablet in der
        Hand.
        """

        if self.selected_plan_id is None:
            self.status_label.text = "Kein Schichtplan gewählt."
            self.status_label.color = theme.TEXT_SECONDARY
            return

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        plan = self.db.get_shift_plan(self.selected_plan_id)
        schichten = self.db.get_shifts(self.selected_plan_id)

        if not schichten:
            self.status_label.text = "Dieser Plan ist noch leer."
            self.status_label.color = theme.TEXT_SECONDARY
            return

        workbook = Workbook()
        blatt = workbook.active
        blatt.title = (plan["event_name"][:31] or "Schichtplan")

        fett = Font(bold=True)

        blatt.append((
            f"Schichtplan {plan['event_name']} "
            f"{self.format_date(plan['event_date'])}",
        ))
        blatt["A1"].font = Font(bold=True, size=14)

        blatt.append(())

        blatt.append((
            "Tätigkeit", "von", "bis", "Soll", "Ist", "fehlen", "Helfer",
        ))

        for zelle in blatt[3]:
            zelle.font = fett

        for schicht in schichten:

            namen = ", ".join(
                helfer["name"]
                for helfer in self.db.get_shift_helpers(schicht["id"])
            )

            fehlen = max(0, schicht["needed"] - schicht["besetzt"])

            blatt.append((
                schicht["task"],
                schicht["start_time"] or "",
                schicht["end_time"] or "",
                schicht["needed"],
                schicht["besetzt"],
                fehlen if fehlen else "",
                namen,
            ))

        besetzt, plaetze, offen = self.db.get_shift_plan_summary(plan["id"])

        blatt.append(())
        blatt.append((
            f"{besetzt} von {plaetze} Plätzen besetzt"
            + (f", {offen} Schichten brauchen noch Helfer" if offen else ""),
        ))
        blatt[blatt.max_row][0].font = fett

        for spalte, breite in zip("ABCDEFG", (26, 10, 10, 8, 8, 10, 48)):
            blatt.column_dimensions[spalte].width = breite

        for zeile in blatt.iter_rows(min_row=4, min_col=7, max_col=7):
            for zelle in zeile:
                zelle.alignment = Alignment(wrap_text=True, vertical="top")

        blatt.page_setup.orientation = "landscape"
        blatt.page_setup.fitToWidth = 1
        blatt.sheet_properties.pageSetUpPr.fitToPage = True
        blatt.print_title_rows = "3:3"

        sicherer_name = "".join(
            zeichen if zeichen.isalnum() or zeichen in " -_" else "_"
            for zeichen in plan["event_name"]
        ).strip().replace(" ", "_")

        dateiname = (
            f"schichtplan_{sicherer_name or 'plan'}_"
            f"{datetime.now():%Y-%m-%d}.xlsx"
        )

        ziel = storage.export_dir("excel") / dateiname

        workbook.save(ziel)

        self.letzte_ausgabe = ziel

        self.status_label.text = export_hinweis(ziel)
        self.status_label.color = theme.TEXT_SECONDARY

    # =====================================================
    # Screen wird geöffnet
    # =====================================================

    def on_pre_enter(self, *_args):

        self._refresh_plans()
        self._refresh_shifts()

    # =====================================================
    # Hilfsmittel
    # =====================================================


    def teilen_clicked(self):
        """Gibt die zuletzt ausgegebene Datei weiter (siehe teilen.py)."""

        erfolg, meldung = teilen.teilen(self.letzte_ausgabe)

        self.status_label.text = meldung

    @staticmethod
    def format_date(iso_date):

        if not iso_date:
            return ""

        try:
            return datetime.strptime(iso_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        except (TypeError, ValueError):
            return str(iso_date)

    @staticmethod
    def _button(text, callback, background=None, color=None):

        button = Button(
            text=text, background_normal="", background_down="",
            background_color=background or theme.SURFACE,
            color=color or theme.TEXT_PRIMARY,
            font_size="15sp", bold=True,
        )
        button.bind(on_release=lambda *_args: callback())

        return button

    @staticmethod
    def _title(text):

        label = KiGLabel(text=text)
        label.set_font_size(24)
        label.set_bold(True)
        label.set_alignment("left")
        label.set_color(theme.PRIMARY_ORANGE)
        label.size_hint_y = None
        label.height = dp(36)

        return label
