"""Auswertung der über die Kasse abgeschlossenen Verkäufe."""

from datetime import datetime

from kivy.graphics import Color, RoundedRectangle
from kivy.metrics import dp
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.screenmanager import Screen
from kivy.uix.spinner import Spinner
from kivy.uix.widget import Widget

import config
import geldformat
import storage
import teilen
import theme
from database import DatabaseManager
from widgets.common.kig_popup import KiGPopup
from widgets.common.kig_symbol import KiGSymbolButton, KREUZ
from widgets.common.date_picker_popup import DatePickerPopup
from widgets.common.exporthinweis import (
    export_hinweis, hinweisfeld_vorbereiten,
)
from widgets.common.feldausrichtung import links_ausrichten
from widgets.common.filterleiste import Filterleiste
from widgets.common.rounded_panel import RoundedPanel
from widgets.kig_label import KiGLabel
from widgets.statistics.category_pie import CategoryPiePanel


class SaleRow(ButtonBehavior, BoxLayout):
    """Eine auswählbare Zeile der Verkaufstabelle."""

    def __init__(self, sale, selected_callback, **kwargs):
        super().__init__(orientation="horizontal", size_hint_y=None, height=dp(46), **kwargs)
        self.sale = sale
        self.selected_callback = selected_callback
        self.selected = False

        # Stornos stehen mit negativer Menge in denselben Tabellen. Die
        # Menge selbst wird in dieser Tabelle nicht angezeigt - ohne
        # eigene Kennzeichnung wäre eine Stornozeile also nicht von
        # einem Verkauf zu unterscheiden.
        self.ist_storno = (sale["quantity"] or 0) < 0

        # Grundfarbe der Zeile merken: Beim Abwählen muss die
        # Storno-Einfärbung zurückkommen und nicht das normale Weiß.
        self.grundfarbe = theme.STORNO_ROW if self.ist_storno else theme.CARD

        with self.canvas.before:
            self._color = Color(*self.grundfarbe)
            self._background = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(6)])
        self.bind(pos=self._refresh_canvas, size=self._refresh_canvas)

        artikel = sale["article_name"]
        if self.ist_storno:
            artikel = f"Storno: {artikel}"

        if theme.is_narrow():

            values = (
                StatisticsScreen.format_date(sale["business_date"]),
                artikel,
                StatisticsScreen.money(sale["unit_price"]),
                StatisticsScreen.money(sale["profit"]),
            )

        else:

            values = (
                sale["event_name"],
                StatisticsScreen.format_date(sale["business_date"]),
                sale["category_name"],
                artikel,
                StatisticsScreen.money(sale["unit_price"]),
                StatisticsScreen.money(sale["purchase_price"]),
                StatisticsScreen.money(sale["profit"]),
            )

        # Dieselbe Aufteilung wie die Kopfzeile - beide kommen aus
        # StatisticsScreen.spalten().
        widths = StatisticsScreen.spalten()[1]

        for spalte, (value, width) in enumerate(zip(values, widths)):
            # Nur der Gewinn wird rot - der Betrag selbst bleibt lesbar,
            # das Vorzeichen macht die Richtung deutlich.
            farbe = (
                theme.ERROR
                if self.ist_storno and spalte == len(values) - 1
                else theme.TEXT_PRIMARY
            )
            label = Label(
                text=str(value), color=farbe, font_size="13sp",
                halign="left", valign="middle", text_size=(None, dp(46)),
                size_hint_x=width,
            )
            self.add_widget(label)

    def _refresh_canvas(self, *_args):
        self._background.pos = self.pos
        self._background.size = self.size

    def on_release(self):
        self.selected = not self.selected
        self._color.rgba = (
            theme.PRIMARY_ORANGE_LIGHT if self.selected else self.grundfarbe
        )
        self.selected_callback(self, self.selected)


class BarGraphic(Widget):
    """Schlichtes horizontales Balkendiagramm ohne externe Bibliothek."""

    def __init__(self, ratio=0, **kwargs):
        super().__init__(**kwargs)
        self.ratio = ratio
        with self.canvas:
            Color(*theme.PRIMARY_ORANGE)
            self._bar = RoundedRectangle(radius=[dp(5)])
        self.bind(pos=self._draw, size=self._draw)

    def _draw(self, *_args):
        self._bar.pos = self.pos
        self._bar.size = (max(0, self.width * self.ratio), self.height)


class StatisticsScreen(Screen):
    """Tabellarische Umsatz- und Gewinnübersicht je Verkaufsposition."""

    # Höhe einer Kennzahlenzeile (Einnahmen, Ausgaben, Gewinn)
    TOTAL_ROW_HEIGHT = 34
    NARROW_TOTAL_ROW_HEIGHT = 28

    def __init__(self, revenue_changed_callback=None, **kwargs):
        super().__init__(**kwargs)
        self.db = DatabaseManager()
        self.revenue_changed_callback = revenue_changed_callback
        self.selected_rows = {}

        # Zuletzt ausgegebene Datei - sie haengt am Teilen-Knopf.
        self.letzte_ausgabe = None
        self.event_options = {"Alle Events": None}

        # Im Hochformat steht die Auswertung unter der Verkaufstabelle
        # statt daneben (siehe theme.set_orientation).
        self.hochformat = theme.is_portrait()

        if theme.is_narrow():
            self.TOTAL_ROW_HEIGHT = self.NARROW_TOTAL_ROW_HEIGHT

        root = BoxLayout(
            orientation="vertical" if self.hochformat else "horizontal",
            padding=dp(theme.SCREEN_PADDING),
            spacing=dp(theme.SCREEN_SPACING),
        )
        root.add_widget(self._build_sales_panel())
        root.add_widget(self._build_summary_panel())
        self.add_widget(root)

    def _build_sales_panel(self):
        schmal = theme.is_narrow()

        # Auf dem Telefon ruecken die Zeilen enger zusammen: Sonst
        # brauchen Ueberschrift, Filter, Kopfzeile und die beiden
        # Loeschknoepfe zusammen mehr Hoehe, als die Karte hat - und
        # die Ueberschrift wurde oben aus ihr hinausgedrueckt.
        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.SPACE_S if schmal else theme.CARD_PADDING),
            spacing=dp(theme.SPACE_XS if schmal else theme.CARD_SPACING),
            size_hint=(
                (1, 0.56 if schmal else 0.62)
                if self.hochformat else (0.64, 1)
            ),
        )

        title = KiGLabel(text="Verkäufe")
        title.set_font_size(26)
        title.set_bold(True)
        title.set_alignment("left")
        title.set_color(theme.PRIMARY_ORANGE)
        title.size_hint_y = None
        title.height = dp(38)
        panel.add_widget(title)

        # Auf dem Telefon passen Ereignisauswahl, zwei Datumsfelder und
        # "Aktualisieren" nicht in eine Zeile - dort brechen sie um.
        schmal = theme.is_narrow()

        filters = BoxLayout(
            orientation="vertical" if schmal else "horizontal",
            size_hint_y=None,
            height=dp(44 * 2 + theme.ROW_SPACING) if schmal else dp(48),
            spacing=dp(theme.ROW_SPACING),
        )

        obere = BoxLayout(spacing=dp(theme.ROW_SPACING)) if schmal else filters
        untere = BoxLayout(spacing=dp(theme.ROW_SPACING)) if schmal else filters

        if schmal:
            filters.add_widget(obere)
            filters.add_widget(untere)

        self.event_filter = Spinner(
            text="Alle Events", values=("Alle Events",),
            font_size="15sp", size_hint_x=1.15,
        )
        links_ausrichten(self.event_filter)
        self.event_filter.bind(text=lambda *_args: self._filter_geaendert())
        obere.add_widget(self.event_filter)

        self.date_from_value = None
        self.date_to_value = None

        untere.add_widget(self._build_date_filter(
            "Von", lambda: self.open_date_picker("from"), lambda: self.clear_date_filter("from"),
        ))
        untere.add_widget(self._build_date_filter(
            "Bis", lambda: self.open_date_picker("to"), lambda: self.clear_date_filter("to"),
        ))

        aktualisieren = self._button("Aktualisieren", self.refresh, width=dp(130))

        if schmal:
            # In der oberen Zeile neben der Ereignisauswahl - unten
            # brauchen die beiden Datumsfelder den ganzen Platz.
            aktualisieren.width = dp(120)
            obere.add_widget(aktualisieren)
        else:
            filters.add_widget(aktualisieren)

        # Ereignis und Zeitraum sind der Filter dieses Bildschirms -
        # sie stehen jetzt unten in der Leiste. Die Karte darueber
        # gehoert damit ganz der Tabelle.
        self.filterleiste = Filterleiste(
            inhalt=filters,
            titel="Auswahl",
            zusammenfassung=self._filter_text,
            inhalt_hoehe=(
                filters.height / dp(1) + 2 * theme.SPACE_XS
            ),
        )

        actions_top = BoxLayout(
            size_hint_y=None, height=dp(36 if schmal else 40),
            spacing=dp(theme.ROW_SPACING),
        )

        if schmal:
            # 170 + 110 dp neben einem Platzhalter passen auf ein
            # Telefon nicht - dort teilen sich beide, was da ist.
            actions_top.add_widget(
                self._button("Export", self.export_excel)
            )
            actions_top.add_widget(
                self._button("Teilen", self.teilen_clicked)
            )

        else:
            actions_top.add_widget(Widget())
            actions_top.add_widget(self._button("Excel exportieren", self.export_excel, width=dp(170)))
            actions_top.add_widget(self._button("Teilen", self.teilen_clicked, width=dp(110)))

        panel.add_widget(actions_top)

        # Eigene Zeile unter den Knoepfen: Der Hinweis nennt den Ordner
        # mit, und ein vollstaendiger Pfad braucht die ganze Breite
        # (siehe widgets/common/exporthinweis.py).
        self.export_status = Label(
            text="", color=theme.TEXT_SECONDARY, font_size="13sp",
            halign="left", valign="middle",
        )

        hinweisfeld_vorbereiten(self.export_status, 0)

        panel.add_widget(self.export_status)

        panel.add_widget(self._build_repair_hint())

        # Bewusst OHNE Abstand zwischen den Spalten: die Kopfzeile muss
        # exakt dieselbe Spaltenaufteilung haben wie SaleRow (dort
        # ebenfalls kein spacing), sonst stehen Überschrift und Wert
        # nicht mehr übereinander. Beide Stellen also nur gemeinsam ändern.
        header = BoxLayout(size_hint_y=None, height=dp(34), spacing=0)
        for title_text, width in zip(*self.spalten()):
            label = Label(
                text=title_text, bold=True, color=theme.TEXT_PRIMARY,
                font_size="13sp", halign="left", valign="middle",
                text_size=(None, dp(34)), size_hint_x=width,
            )
            header.add_widget(label)
        panel.add_widget(header)

        self.sales_rows = BoxLayout(orientation="vertical", spacing=dp(theme.SPACE_XS), size_hint_y=None)
        self.sales_rows.bind(minimum_height=self.sales_rows.setter("height"))
        scroll = ScrollView(do_scroll_x=False)
        scroll.add_widget(self.sales_rows)
        panel.add_widget(scroll)

        actions = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(theme.ROW_SPACING))
        actions.add_widget(self._button("Ausgewählte löschen", self.delete_selected))
        actions.add_widget(self._button("Zeitraum löschen", self.delete_period))
        panel.add_widget(actions)
        return panel

    @staticmethod
    def spalten():
        """Ueberschriften und Breiten der Verkaufstabelle.

        Auf dem Telefon vier statt sieben: Sieben Ueberschriften ergaben
        auf 412 dp einen einzigen Streifen, in dem "VerkaufEinkaufGewinn"
        uebereinanderlag. Event, Kategorie und Einkauf entfallen dort -
        wer sie braucht, sieht sie in der Ausgabe nach Excel.
        """

        if theme.is_narrow():
            return (
                ("Datum", "Artikel", "Verkauf", "Gewinn"),
                (0.22, 0.40, 0.19, 0.19),
            )

        return (
            ("Event", "Datum", "Kategorie", "Artikel",
             "Verkauf", "Einkauf", "Gewinn"),
            (0.20, 0.12, 0.15, 0.20, 0.11, 0.11, 0.11),
        )

    # =====================================================
    # Fehlende Einkaufspreise nachtragen
    # =====================================================

    def _build_repair_hint(self):
        """Zeile, die auf Rezeptverkäufe ohne Einkaufspreis hinweist.

        Solche Positionen entstehen, wenn beim Verkauf die Kosten der
        Zutaten nicht bestimmbar waren (z. B. Flasche ohne
        Wareneingang): Gebucht wurde dann 0,00, der Gewinn steht damit
        zu hoch. Die Zeile bleibt unsichtbar, solange alles stimmt.
        """

        self.repair_row = BoxLayout(
            size_hint_y=None, height=0, opacity=0,
            spacing=dp(theme.ROW_SPACING),
        )

        self.repair_label = Label(
            text="", color=theme.ERROR, font_size="13sp",
            halign="left", valign="middle",
        )
        self.repair_label.bind(
            size=lambda instance, value: setattr(instance, "text_size", value)
        )

        self.repair_row.add_widget(self.repair_label)
        self.repair_row.add_widget(
            self._button("Einkaufspreise nachtragen", self.repair_costs, width=dp(220))
        )

        return self.repair_row

    def _refresh_repair_hint(self, date_from, date_to, event_id):

        offen = self.db.count_missing_recipe_costs(date_from, date_to, event_id)

        if not offen:
            self.repair_row.height = 0
            self.repair_row.opacity = 0
            self.repair_row.disabled = True
            return

        self.repair_label.text = (
            f"{offen} {'Verkauf' if offen == 1 else 'Verkäufe'} von "
            "Rezeptartikeln ohne Einkaufspreis - der Gewinn steht zu hoch."
        )

        self.repair_row.height = dp(40)
        self.repair_row.opacity = 1
        self.repair_row.disabled = False

    def repair_costs(self):

        date_from, date_to = self._period()
        event_id = self.event_options.get(self.event_filter.text)

        offen = self.db.count_missing_recipe_costs(date_from, date_to, event_id)

        if not offen:
            return

        self._confirm(
            f"Bei {offen} Verkaufsposition(en) den heute gültigen "
            "Rezeptpreis als Einkaufspreis nachtragen?",
            lambda: self._repair_costs_confirmed(date_from, date_to, event_id),
            confirm_text="Nachtragen",
        )

    def _repair_costs_confirmed(self, date_from, date_to, event_id):

        nachgetragen = self.db.repair_recipe_costs(date_from, date_to, event_id)

        self.refresh()

        self.export_status.text = (
            f"{nachgetragen} Einkaufspreis(e) nachgetragen."
            if nachgetragen else
            "Kein Preis nachtragbar - bitte zuerst den Wareneingang "
            "der Zutaten mit Preis buchen."
        )

    def _build_summary_panel(self):
        """Die rechte Spalte: oben die Zahlen des Zeitraums, unten die
        Rangliste der Artikel.

        Beide Karten zeigen ausschließlich den links eingestellten
        Zeitraum (und das gewählte Event) - Filter und Auswertung
        gehören zusammen, sonst stünden Tabelle und Diagramm für
        verschiedene Zeiträume nebeneinander.

        Hochformat: Die beiden Karten stehen unter der Tabelle
        nebeneinander - untereinander bliebe von beiden nur noch ein
        Streifen übrig.
        """

        # Auf dem Telefon aber wieder untereinander: 180 dp je Karte
        # reichen nicht einmal fuer die Ueberschrift
        # "Gesamtverkaufszahlen".
        nebeneinander = self.hochformat and not theme.is_narrow()

        panel = BoxLayout(
            orientation="horizontal" if nebeneinander else "vertical",
            spacing=dp(theme.SCREEN_SPACING),
            size_hint=(
                (1, 0.44 if theme.is_narrow() else 0.38)
                if self.hochformat else (0.36, 1)
            ),
        )

        panel.add_widget(self._build_totals_panel())
        panel.add_widget(self._build_top_panel())

        return panel

    def _build_totals_panel(self):
        """Gesamteinnahmen, Gesamtausgaben, Gewinn - und darunter, wie
        sich die Einnahmen auf die Kategorien verteilen."""

        schmal = theme.is_narrow()

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.SPACE_S if schmal else theme.CARD_PADDING),
            spacing=dp(theme.SPACE_XS if schmal else theme.CARD_SPACING),
            size_hint=(
                (0.5, 1) if self.hochformat and not schmal
                else (1, 0.58)
            ),
        )

        panel.add_widget(self._title("Gesamtverkaufszahlen"))

        self.total_labels = {}

        kennzahlen = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            height=dp(3 * self.TOTAL_ROW_HEIGHT),
        )

        for schluessel, beschriftung, farbe in (
            ("revenue", "Einnahmen", theme.TEXT_PRIMARY),
            ("expenses", "Ausgaben", theme.TEXT_PRIMARY),
            ("profit", "Gewinn", theme.PRIMARY_ORANGE),
        ):
            kennzahlen.add_widget(self._total_row(schluessel, beschriftung, farbe))

        panel.add_widget(kennzahlen)

        self.period_label = Label(
            text="", color=theme.TEXT_SECONDARY, font_size="13sp",
            size_hint_y=None, height=dp(22),
            halign="left", valign="middle",
        )
        self.period_label.bind(
            size=lambda instance, value: setattr(instance, "text_size", value)
        )
        panel.add_widget(self.period_label)

        self.category_pie = CategoryPiePanel()

        # Auf dem Telefon bleibt das Tortendiagramm weg: In den rund
        # 150 dp, die der Karte dort bleiben, ist es nicht zu lesen -
        # und seine Legende legte sich ueber die Zeile darueber. Die
        # drei Zahlen sagen auf dem kleinen Schirm genug.
        if not theme.is_narrow():
            panel.add_widget(self.category_pie)

        return panel

    def _total_row(self, schluessel, beschriftung, farbe):

        zeile = BoxLayout(size_hint_y=None, height=dp(self.TOTAL_ROW_HEIGHT))

        zeile.add_widget(Label(
            text=beschriftung, color=theme.TEXT_SECONDARY, font_size="15sp",
            halign="left", valign="middle", size_hint_x=0.5,
            text_size=(None, dp(self.TOTAL_ROW_HEIGHT)),
        ))

        wert = Label(
            text=self.money(0), color=farbe, font_size="19sp", bold=True,
            halign="right", valign="middle", size_hint_x=0.5,
            text_size=(None, dp(self.TOTAL_ROW_HEIGHT)),
        )

        self.total_labels[schluessel] = wert
        zeile.add_widget(wert)

        return zeile

    def _build_top_panel(self):
        """Balkendiagramm der meistverkauften Artikel."""

        panel = RoundedPanel(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
            size_hint=(
                (0.5, 1) if self.hochformat and not theme.is_narrow()
                else (1, 0.42)
            ),
        )

        panel.add_widget(self._title("Top-Artikel"))

        # Fünf Zeilen zu 34 dp plus Abstand brauchen mehr Platz, als die
        # Karte je nach Fenstergröße hergibt - ohne ScrollView zeichneten
        # die untersten Zeilen über den Kartenrand hinaus (dieselbe
        # Lösung wie bei der Tabelle links).
        self.top_rows = BoxLayout(
            orientation="vertical",
            spacing=dp(theme.CARD_SPACING),
            size_hint_y=None,
        )
        self.top_rows.bind(minimum_height=self.top_rows.setter("height"))

        top_scroll = ScrollView(do_scroll_x=False)
        top_scroll.add_widget(self.top_rows)
        panel.add_widget(top_scroll)

        return panel

    @staticmethod
    def _button(text, callback, width=None):
        button = Button(
            text=text, background_normal="", background_down="",
            background_color=theme.SURFACE, color=theme.TEXT_PRIMARY,
            font_size="15sp", bold=True, size_hint_x=None if width else 1,
        )
        if width:
            button.width = width
        button.bind(on_release=lambda *_args: callback())
        return button

    @staticmethod
    def _title(text):
        schmal = theme.is_narrow()

        label = KiGLabel(text=text)
        label.set_font_size(17 if schmal else 22)
        label.set_bold(True)
        label.set_alignment("left")
        label.set_color(theme.PRIMARY_ORANGE)
        label.size_hint_y = None
        label.height = dp(26 if schmal else 34)
        return label


    def teilen_clicked(self):
        """Gibt die zuletzt ausgegebene Datei weiter (siehe teilen.py)."""

        erfolg, meldung = teilen.teilen(self.letzte_ausgabe)

        self.export_status.text = meldung

    @staticmethod
    def money(value):
        return geldformat.geld(value)

    @staticmethod
    def format_date(iso_date):
        try:
            return datetime.strptime(iso_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        except (TypeError, ValueError):
            return str(iso_date or "-")

    def on_pre_enter(self, *_args):
        self._populate_event_filter()
        self.refresh()

    def _populate_event_filter(self):
        """Bietet alle im Kalender gepflegten Events als optionalen Filter an."""

        current_selection = self.event_filter.text
        self.event_options = {"Alle Events": None}

        for event in self.db.get_events():
            if event["entry_type"] != "EVENT":
                continue
            label = f"{event['name']} ({self.format_date(event['start_date'])})"
            self.event_options[label] = event["id"]

        self.event_filter.values = tuple(self.event_options)
        self.event_filter.text = (
            current_selection if current_selection in self.event_options
            else "Alle Events"
        )

    def _period(self):
        return self.date_from_value, self.date_to_value

    def _filter_geaendert(self):

        self.refresh()

        if getattr(self, "filterleiste", None) is not None:
            self.filterleiste.aktualisieren()

    def _filter_text(self):
        """Was in der zugeklappten Filterleiste steht."""

        teile = [self.event_filter.text]

        von = self.date_from_value
        bis = self.date_to_value

        if von and bis:
            teile.append(f"{self.format_date(von)} - {self.format_date(bis)}")
        elif von:
            teile.append(f"ab {self.format_date(von)}")
        elif bis:
            teile.append(f"bis {self.format_date(bis)}")
        else:
            teile.append("ganzer Zeitraum")

        return "   ·   ".join(teile)

    # =====================================================
    # Zeitraum-Auswahl (Kalender-Dropdown)
    # =====================================================

    def _build_date_filter(self, label_prefix, on_pick, on_clear):
        """Ein Datumsfeld als Button (öffnet den Kalender) + Löschen-Knopf,
        anstatt das Datum von Hand eintippen zu müssen."""

        box = BoxLayout(spacing=dp(theme.LABEL_SPACING))

        button = self._button(f"{label_prefix}: alle", on_pick)

        # Datumsfeld, kein Aktionsknopf: linksbuendig.
        links_ausrichten(button)

        box.add_widget(button)

        clear_button = KiGSymbolButton(
            symbol=KREUZ, symbol_color=theme.TEXT_SECONDARY,
            size_hint_x=None, width=dp(36),
            background_color=theme.SURFACE,
        )
        clear_button.bind(on_release=lambda *_args: on_clear())
        box.add_widget(clear_button)

        if label_prefix == "Von":
            self.date_from_button = button
        else:
            self.date_to_button = button

        return box

    def open_date_picker(self, which):

        current = self.date_from_value if which == "from" else self.date_to_value

        DatePickerPopup(
            title="Von: Startdatum" if which == "from" else "Bis: Enddatum",
            initial_date=current,
            on_select=lambda iso_date: self.date_picked(which, iso_date),
        ).open()

    def date_picked(self, which, iso_date):

        label_prefix = "Von" if which == "from" else "Bis"
        button = self.date_from_button if which == "from" else self.date_to_button

        if which == "from":
            self.date_from_value = iso_date
        else:
            self.date_to_value = iso_date

        button.text = f"{label_prefix}: {self.format_date(iso_date)}"
        self.refresh()

        self.filterleiste.aktualisieren()

    def clear_date_filter(self, which):

        label_prefix = "Von" if which == "from" else "Bis"
        button = self.date_from_button if which == "from" else self.date_to_button

        if which == "from":
            self.date_from_value = None
        else:
            self.date_to_value = None

        button.text = f"{label_prefix}: alle"
        self.refresh()

        self.filterleiste.aktualisieren()

    def refresh(self):
        date_from, date_to = self._period()
        event_id = self.event_options.get(self.event_filter.text)
        self.selected_rows.clear()
        self.sales_rows.clear_widgets()

        rows = self.db.get_statistic_sale_items(date_from, date_to, event_id)
        if not rows:
            self.sales_rows.add_widget(self._empty_label("Keine Verkäufe im gewählten Zeitraum."))
        else:
            for row in rows:
                self.sales_rows.add_widget(SaleRow(row, self._row_selected))

        self._refresh_totals(date_from, date_to, event_id)
        self._refresh_top_articles(date_from, date_to, event_id)
        self._refresh_repair_hint(date_from, date_to, event_id)

    def _row_selected(self, row, selected):
        row_key = row.sale["row_key"]
        if selected:
            self.selected_rows[row_key] = row.sale["sale_item_id"]
        else:
            self.selected_rows.pop(row_key, None)

    def _refresh_totals(self, date_from, date_to, event_id):
        """Kennzahlen und Kreisdiagramm - beide für denselben Zeitraum
        wie die Tabelle links."""

        kennzahlen = self.db.get_period_totals(date_from, date_to, event_id)

        for schluessel, label in self.total_labels.items():
            label.text = self.money(kennzahlen[schluessel])

        self.period_label.text = self._period_text(kennzahlen)

        self.category_pie.set_data(
            self.db.get_category_revenues(date_from, date_to, event_id)
        )

    def _period_text(self, kennzahlen):
        """Eine Zeile, die sagt, worauf sich die Zahlen beziehen.

        Ohne sie ließe sich am Diagramm nicht ablesen, ob gerade ein
        Zeitraum eingegrenzt ist oder alles gezeigt wird.
        """

        date_from, date_to = self._period()

        if date_from and date_to:
            zeitraum = f"{self.format_date(date_from)} - {self.format_date(date_to)}"
        elif date_from:
            zeitraum = f"ab {self.format_date(date_from)}"
        elif date_to:
            zeitraum = f"bis {self.format_date(date_to)}"
        else:
            zeitraum = "gesamter Zeitraum"

        if self.event_filter.text != "Alle Events":
            zeitraum = f"{self.event_filter.text} | {zeitraum}"

        bons = kennzahlen["receipts"]

        return (
            f"{zeitraum} | {kennzahlen['quantity']} Einheiten auf "
            f"{bons} {'Bon' if bons == 1 else 'Bons'}"
        )

    def _refresh_top_articles(self, date_from, date_to, event_id):
        self.top_rows.clear_widgets()
        top_articles = self.db.get_top_selling_articles(date_from, date_to, event_id)
        if not top_articles:
            self.top_rows.add_widget(self._empty_label("Noch keine Verkaufsdaten."))
            return

        maximum = top_articles[0][1]
        for index, (name, quantity) in enumerate(top_articles, start=1):
            # Rechts etwas Luft: Dort liegt der Rollbalken der Liste,
            # und die Menge stand sonst halb darunter.
            row = BoxLayout(
                size_hint_y=None, height=dp(34), spacing=dp(theme.ROW_SPACING),
                padding=[0, 0, dp(12), 0],
            )
            row.add_widget(Label(
                text=f"{index}. {name}", color=theme.TEXT_PRIMARY, font_size="14sp",
                halign="left", valign="middle", text_size=(dp(105), dp(34)), size_hint_x=None, width=dp(105),
                shorten=True, shorten_from="right",
            ))
            row.add_widget(BarGraphic(ratio=quantity / maximum if maximum else 0))
            row.add_widget(Label(
                text=str(quantity), color=theme.TEXT_PRIMARY, font_size="14sp",
                size_hint_x=None, width=dp(36), halign="right", valign="middle", text_size=(dp(36), dp(34)),
            ))
            self.top_rows.add_widget(row)

    @staticmethod
    def _empty_label(text):
        return Label(
            text=text, color=theme.TEXT_SECONDARY, font_size="14sp", size_hint_y=None,
            height=dp(42), halign="left", valign="middle", text_size=(None, dp(42)),
        )

    # =====================================================
    # Excel-Export
    # =====================================================

    def export_excel(self):
        """Exportiert die aktuell gefilterten Verkaufszahlen inklusive
        Diagrammen als Excel-Arbeitsmappe (zwei Blätter: Rohdaten und
        Auswertung mit Balkendiagrammen)."""

        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font

        date_from, date_to = self._period()
        event_id = self.event_options.get(self.event_filter.text)

        sale_items = self.db.get_statistic_sale_items(date_from, date_to, event_id)
        top_articles = self.db.get_top_selling_articles(date_from, date_to, event_id, limit=5)
        revenues = self.db.get_article_revenues(date_from, date_to, event_id)

        if not sale_items:
            self.export_status.text = "Keine Verkäufe im gewählten Zeitraum zum Exportieren."
            return

        workbook = Workbook()
        bold = Font(bold=True)

        # -------------------------------------------------
        # Blatt 1: Verkäufe (Rohdaten)
        # -------------------------------------------------

        sales_sheet = workbook.active
        sales_sheet.title = "Verkäufe"

        headers = ("Event", "Datum", "Kategorie", "Artikel", "Menge", "Verkauf", "Einkauf", "Gewinn")
        sales_sheet.append(headers)
        for cell in sales_sheet[1]:
            cell.font = bold

        for row in sale_items:
            sales_sheet.append((
                row["event_name"], self.format_date(row["business_date"]), row["category_name"],
                row["article_name"], row["quantity"], row["unit_price"], row["purchase_price"],
                row["profit"],
            ))

        for column_cells in sales_sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sales_sheet.column_dimensions[column_cells[0].column_letter].width = max(10, width + 2)

        # -------------------------------------------------
        # Blatt 2: Auswertung (Top 5 + Gesamtverkaufszahlen)
        # -------------------------------------------------

        summary_sheet = workbook.create_sheet("Auswertung")

        summary_sheet.append(("Top 5 Positionen (Menge)",))
        summary_sheet["A1"].font = bold
        summary_sheet.append(("Artikel", "Menge"))
        for cell in summary_sheet[2]:
            cell.font = bold
        top_start_row = 3
        for name, quantity in top_articles:
            summary_sheet.append((name, quantity))
        top_end_row = top_start_row + len(top_articles) - 1

        if top_articles:
            top_chart = BarChart()
            top_chart.title = "Top 5 Positionen"
            top_chart.y_axis.title = "Menge"
            data = Reference(summary_sheet, min_col=2, min_row=2, max_row=top_end_row)
            categories = Reference(summary_sheet, min_col=1, min_row=top_start_row, max_row=top_end_row)
            top_chart.add_data(data, titles_from_data=True)
            top_chart.set_categories(categories)
            top_chart.width, top_chart.height = 16, 9
            summary_sheet.add_chart(top_chart, "D2")

        revenue_start_row = top_end_row + 3
        summary_sheet.cell(row=revenue_start_row, column=1, value="Gesamtverkaufszahlen (Einnahmen)").font = bold
        header_row = revenue_start_row + 1
        summary_sheet.cell(row=header_row, column=1, value="Artikel").font = bold
        summary_sheet.cell(row=header_row, column=2, value="Einnahmen").font = bold

        revenue_data_start = header_row + 1
        for offset, (name, revenue) in enumerate(revenues):
            summary_sheet.cell(row=revenue_data_start + offset, column=1, value=name)
            summary_sheet.cell(row=revenue_data_start + offset, column=2, value=round(revenue, 2))
        revenue_data_end = revenue_data_start + len(revenues) - 1

        if revenues:
            revenue_chart = BarChart()
            revenue_chart.title = "Gesamtverkaufszahlen"
            revenue_chart.y_axis.title = "Einnahmen (€)"
            data = Reference(summary_sheet, min_col=2, min_row=header_row, max_row=revenue_data_end)
            categories = Reference(summary_sheet, min_col=1, min_row=revenue_data_start, max_row=revenue_data_end)
            revenue_chart.add_data(data, titles_from_data=True)
            revenue_chart.set_categories(categories)
            revenue_chart.width, revenue_chart.height = 16, 9
            summary_sheet.add_chart(revenue_chart, f"D{revenue_start_row}")

        summary_sheet.column_dimensions["A"].width = 28
        summary_sheet.column_dimensions["B"].width = 14

        # -------------------------------------------------
        # Speichern
        # -------------------------------------------------

        filename = datetime.now().strftime("statistik_%Y-%m-%d_%H-%M.xlsx")
        export_path = storage.export_dir("excel") / filename
        workbook.save(export_path)

        self.letzte_ausgabe = export_path

        self.export_status.text = export_hinweis(export_path)

    def delete_selected(self):
        if not self.selected_rows:
            return
        self._confirm(
            f"{len(self.selected_rows)} ausgewählte Verkaufsposition(en) löschen?",
            lambda: self._delete_selected_confirmed(),
        )

    def _delete_selected_confirmed(self):
        units_per_sale_item = {}
        for sale_item_id in self.selected_rows.values():
            units_per_sale_item[sale_item_id] = (
                units_per_sale_item.get(sale_item_id, 0) + 1
            )
        for sale_item_id, quantity in units_per_sale_item.items():
            self.db.delete_sale_units(sale_item_id, quantity)
        self.refresh()
        self._notify_revenue_changed()

    def delete_period(self):
        date_from, date_to = self._period()
        if not date_from or not date_to or date_from > date_to:
            self._confirm("Bitte zuerst einen gültigen Zeitraum eingeben.", None)
            return
        self._confirm(
            "Alle Verkäufe im gewählten Zeitraum löschen?",
            lambda: self._delete_period_confirmed(date_from, date_to),
        )

    def _delete_period_confirmed(self, date_from, date_to):
        self.db.delete_sales_in_period(date_from, date_to)
        self.refresh()
        self._notify_revenue_changed()

    def _notify_revenue_changed(self):
        """Meldet Änderungen an den Verkaufsdaten, damit der Header-Tagesumsatz

        (der nur die heutigen Verkäufe zeigt) unmittelbar aktualisiert wird,
        falls gelöschte Positionen den heutigen Geschäftstag betreffen.
        """
        if callable(self.revenue_changed_callback):
            self.revenue_changed_callback()

    def _confirm(self, message, callback, confirm_text="Löschen"):
        content = BoxLayout(
            orientation="vertical",
            padding=dp(theme.CARD_PADDING),
            spacing=dp(theme.CARD_SPACING),
        )
        content.add_widget(Label(text=message, color=theme.TEXT_PRIMARY, font_size="16sp"))
        buttons = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(theme.ROW_SPACING))
        popup = KiGPopup(title="Verkaufsdaten", content=content, size_hint=(0.45, None), height=dp(190), auto_dismiss=False)
        cancel = self._button("Abbrechen", popup.dismiss)
        buttons.add_widget(cancel)
        if callback:
            confirm = self._button(confirm_text, lambda: (popup.dismiss(), callback()))
            buttons.add_widget(confirm)
        content.add_widget(buttons)
        popup.open()
