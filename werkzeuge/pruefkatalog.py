"""
=========================================================
KiG POS
=========================================================

Datei:
    werkzeuge/pruefkatalog.py

Beschreibung:
    Erzeugt eine Excel-Mappe zum Durchtesten der Anwendung -
    Funktion für Funktion, Gerät für Gerät.

    Aufruf:
        .venv\\Scripts\\python.exe werkzeuge\\pruefkatalog.py
        .venv\\Scripts\\python.exe werkzeuge\\pruefkatalog.py <zieldatei>

    Aufbau der Mappe:

        Anleitung    wie die Liste zu benutzen ist
        Übersicht    zählt je Bereich zusammen, was geprüft ist
        je Bereich   ein Blatt mit den Funktionen darin

    Je Zeile eine Funktion, je Gerät eine Spalte. Wo eine
    Funktion auf einem Gerät nicht vorkommt, steht ein
    Strich - dort ist nichts zu prüfen, und niemand soll
    suchen (der PDF-Export gibt es nur am Rechner, die
    Dateiauswahl des Systems nur auf Android, Querformat
    nicht auf dem Telefon).

    Die Liste ist aus dem Programm abgeleitet: Sie folgt
    den Themen des Handbuchs (widgets/userguide/content.py)
    und ergänzt, was dort noch nicht steht - Übertragung
    zwischen Geräten, Sperren auf dem Nebengerät und die
    Handy-Ansicht.

Version:
    1.0.0
=========================================================
"""

import sys

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJEKT = Path(__file__).resolve().parent.parent

sys.path.insert(0, str(PROJEKT))

import config


# =========================================================
# Geräte und Zustände
# =========================================================

GERAETE = ("Rechner", "Tablet", "Handy")

ZUSTAENDE = ("offen", "OK", "Fehler", "Teilweise", "entfällt")

# Kürzel für die Spalten je Gerät: Wo eine Funktion nicht vorkommt,
# steht von vornherein "entfällt".
NICHT = "entfällt"


# =========================================================
# Was beim Neuerzeugen erhalten bleibt
# =========================================================
#
# Der Katalog waechst mit dem Programm - eingetragene Ergebnisse und
# Kommentare duerfen dabei nicht verlorengehen. Beim Erzeugen wird
# eine vorhandene Mappe deshalb gelesen und ihr Stand uebernommen.
#
# Zwei Sonderfaelle:
#
#   UMBENANNT   Die Funktion heisst jetzt anders. Der Kommentar zieht
#               mit, das Ergebnis nicht - die Funktion ist eine
#               andere geworden.
#
#   GEAENDERT   Die Funktion heisst gleich, verhaelt sich aber
#               anders. Auch hier bleibt der Kommentar, das Ergebnis
#               geht auf "offen": Was sich geaendert hat, will neu
#               angesehen werden.

UMBENANNT = {
    ("Grundlagen", "Fußzeile: Version und Beenden"):
        ("Grundlagen", "Version, Build und Beenden in den Einstellungen"),

    ("Grundlagen", "Kopfzeile: Datum, Uhrzeit, Tagesumsatz"):
        ("Grundlagen", "Kopfzeile: Tagesumsatz"),

    ("Darstellung", "Kopf- und Fußzeile schlank"):
        ("Darstellung", "Kopfzeile schlank"),
}

GEAENDERT = {
    ("Grundlagen", "Kopfzeile: Veranstaltung des Tages"),
    ("Kasse", "Warenkorb als Zeile"),
    ("Kasse", "Warenkorb aufklappen"),
    ("Artikelverwaltung", "Übersicht"),
    ("Artikelverwaltung", "Kategorie anlegen"),
    ("Artikelverwaltung", "Kategorie ändern und löschen"),
    ("Kassenbuch", "Zeitraum wählen"),
    ("Statistik", "Nach Event filtern"),
    ("Statistik", "Nach Zeitraum filtern"),
}


FARBEN = {
    "kopf": "F44611",
    "bereich": "FDE8E1",
    "ok": "C6EFCE",
    "fehler": "FFC7CE",
    "teilweise": "FFEB9C",
    "entfaellt": "E7E7E7",
}


# =========================================================
# Der Katalog
# =========================================================
#
# (Funktion, Wo zu finden, Erwartetes Verhalten, Geräte)
#
# Geräte: Zeichenkette aus R (Rechner), T (Tablet), H (Handy).
# Was fehlt, entfällt dort.

KATALOG = [

    ("Grundlagen", [

        ("Programm startet",
         "Programmsymbol / Startseite",
         "Startbild erscheint, danach die Startseite mit den Kacheln.",
         "RTH"),

        ("Leere Datenbank beim ersten Start",
         "Erster Start auf einem neuen Gerät",
         "Sechs Kategorien, keine Artikel, keine Verkäufe.",
         "RTH"),

        ("Automatische Sicherung",
         "Ordner daten/backups",
         "Bei jedem Start entsteht eine Sicherung; die letzten 15 "
         "bleiben liegen.",
         "RTH"),

        ("Kaputte Datenbank",
         "daten/kig.db unlesbar machen (Sicherung vorher!)",
         "Verständlicher Hinweis mit Speicherort und Beenden-Knopf "
         "statt Absturz.",
         "R"),

        ("Kopfzeile: Logo führt heim",
         "Kopfzeile links",
         "Ein Tipp auf das Logo öffnet die Startseite.",
         "RTH"),

        ("Kopfzeile: Tagesumsatz",
         "Kopfzeile rechts",
         "Der Tagesumsatz ändert sich nach einem Verkauf sofort.",
         "RTH"),

        ("Kopfzeile: Datum und Uhrzeit nur am Rechner",
         "Kopfzeile rechts",
         "Am Rechner läuft die Uhr mit. Auf Tablet und Telefon steht "
         "dort nichts dergleichen - das zeigt das Gerät selbst.",
         "RTH"),

        ("Kopfzeile: Veranstaltung des Tages",
         "Kopfzeile Mitte (Event im Kalender anlegen)",
         "Der Name der heutigen Veranstaltung steht dort, sonst "
         "\"KiG POS\". Ein neu angelegtes Event erscheint nach dem "
         "nächsten Bildschirmwechsel.",
         "RT"),

        ("Keine Fußzeile mehr",
         "Alle Bildschirme, unterer Rand",
         "Unten steht nur noch, was zum Bildschirm gehört - Warenkorb "
         "oder Filter. Kein Streifen mit Version und Beenden.",
         "RTH"),

        ("Version, Build und Beenden in den Einstellungen",
         "Einstellungen, Abschnitt \"Programm\"",
         "Version und Buildnummer stehen dort; \"Programm beenden\" "
         "fragt vorher nach und beendet erst dann.",
         "RTH"),

        ("Buildnummer steigt mit jeder Fassung",
         "Einstellungen, Abschnitt \"Programm\"",
         "Nicht mehr fest 0001: Die Nummer ist bei jedem neuen Stand "
         "höher als beim vorigen.",
         "RTH"),

        ("Startbild sitzt richtig",
         "Beim Start, vor der Startseite",
         "Das Logo ist ganz zu sehen und nicht abgeschnitten; Titel, "
         "Untertitel und Wahlspruch überlappen einander nicht.",
         "RTH"),

        ("Startseite: drei Gruppen",
         "Startseite",
         "Operativ, Administrativ, Support - alle neun Kacheln "
         "erreichbar.",
         "RTH"),

        ("Startseite: Gruppen klappbar",
         "Startseite, Überschrift antippen",
         "Die Gruppe klappt zu und wieder auf.",
         "H"),

        ("Startseite: zwei Kacheln je Reihe",
         "Startseite",
         "Zwei Kacheln nebeneinander, nichts abgeschnitten, kein "
         "Scrollen nötig.",
         "H"),

        ("Jede Kachel öffnet ihren Bereich",
         "Startseite, alle neun Kacheln",
         "Jeder Bereich öffnet sich und lässt sich wieder verlassen.",
         "RTH"),

    ]),

    ("Kasse", [

        ("Aufbau",
         "Kasse",
         "Kategorien, Artikel und Warenkorb sind zu sehen.",
         "RT"),

        ("Kategorien als Klappköpfe",
         "Kasse (Telefon)",
         "Kategorien stehen untereinander mit Anzahl; offen ist immer "
         "genau eine.",
         "H"),

        ("Nach Kategorie filtern",
         "Kasse, Kategorie antippen",
         "Nur deren Artikel; zweiter Tipp hebt den Filter auf.",
         "RT"),

        ("Artikel suchen",
         "Kasse, Suchfeld",
         "Treffer erscheinen beim Tippen; das Kreuz setzt zurück.",
         "RTH"),

        ("Suche geht über alle Kategorien",
         "Kasse (Telefon), Suchfeld",
         "Während der Suche treten die Kategorien zurück, Treffer "
         "kommen aus allen.",
         "H"),

        ("Bestand auf der Kachel",
         "Kasse, Artikelkacheln",
         "\"Bestand\" bei normalen Artikeln, \"Verfügbar\" bei "
         "Rezepten; ausverkauft ist grau.",
         "RTH"),

        ("Artikel in den Warenkorb",
         "Kasse, Kachel antippen",
         "Position erscheint im Warenkorb, Summe stimmt.",
         "RTH"),

        ("Menge direkt ändern",
         "Warenkorb, Plus/Minus an der Position",
         "Menge und Summe ändern sich sofort.",
         "RTH"),

        ("Position bearbeiten",
         "Warenkorb, Position antippen, \"Bearbeiten\"",
         "Preis und Menge lassen sich ändern, Löschen und Duplizieren "
         "gehen.",
         "RTH"),

        ("Rezept als Hilfe",
         "Kasse, Mix-Artikel im Warenkorb antippen",
         "Die Zusammensetzung wird eingeblendet.",
         "RTH"),

        ("Warenkorb leeren",
         "Warenkorb, \"Leeren\"",
         "Nach Rückfrage ist der Warenkorb leer.",
         "RTH"),

        ("Warenkorb als Zeile",
         "Kasse im Hochformat, unten",
         "Zugeklappt eine Zeile: Postenzahl, Summe, \"Bezahlen\".",
         "RTH"),

        ("Warenkorb aufklappen",
         "Kasse im Hochformat, auf die Zeile tippen",
         "Er klappt hoch und nimmt den ganzen Bildschirm ein; die "
         "Artikel treten so lange beiseite.",
         "RTH"),

        ("Warenkorb wieder zuklappen",
         "Aufgeklappter Warenkorb, auf \"Warenkorb\" oben tippen",
         "Der Winkel zeigt nach oben; ein Tipp klappt zu, die Artikel "
         "sind wieder da.",
         "RTH"),

        ("Bezahlen aus der zugeklappten Zeile",
         "Kasse im Hochformat, \"Bezahlen\" in der Zeile",
         "Das Zahlungsfenster geht mit der richtigen Summe auf - ohne "
         "vorher aufzuklappen.",
         "RTH"),

        ("Leerer Warenkorb klappt selbst zu",
         "Bezahlen oder Leeren, während er aufgeklappt ist",
         "Er klappt von allein zur Zeile zusammen, statt leer den "
         "Bildschirm zu belegen.",
         "RTH"),

        ("Bezahlen: Nummernblock",
         "Kasse, \"Bezahlen\"",
         "Nummernblock und Infopanel gehen auf; der zu zahlende Betrag "
         "steht im Warenkorb.",
         "RTH"),

        ("Bezahlen: Schnellwahl 5/10/20/50/100",
         "Bezahlen, Schnellwahlfelder",
         "Jeder Tipp legt einen Schein dazu (2x20 = 40); darunter "
         "steht die Aufstellung.",
         "RTH"),

        ("Bezahlen: Rückgeld",
         "Bezahlen",
         "Gegeben und Rückgeld stimmen; Rückgeld wird erst grün, wenn "
         "es reicht.",
         "RTH"),

        ("Verkauf abschließen",
         "Bezahlen, bestätigen",
         "Bon wird gebucht, Warenkorb leer, Bestand sinkt, Tagesumsatz "
         "steigt.",
         "RTH"),

        ("Stornieren",
         "Kasse, \"Storno\"",
         "Artikel antippen, bestätigen: Gegenbuchung mit negativer "
         "Menge, Bestand steigt wieder.",
         "RTH"),

        ("Beträge mit Komma",
         "Überall in der Kasse",
         "2,50 € statt 2.50 - im Warenkorb, auf den Kacheln und beim "
         "Bezahlen.",
         "RTH"),

    ]),

    ("Artikelverwaltung", [

        ("Übersicht",
         "Artikel",
         "Die Artikelliste hat den ganzen Bildschirm; die Kategorien "
         "stehen unten in der Leiste.",
         "RTH"),

        ("Kategorien in der Leiste",
         "Artikel, Leiste unten \"Kategorie\"",
         "Zugeklappt steht dort die gewählte Kategorie (sonst \"Alle "
         "Kategorien\"); ein Tipp klappt die Kategorien hoch.",
         "RTH"),

        ("Nach Kategorie filtern",
         "Artikel, Kategorie in der Leiste antippen",
         "Die Liste zeigt nur deren Artikel, die Leiste nennt sie; "
         "ein zweiter Tipp hebt den Filter auf.",
         "RTH"),

        ("Kategorie anlegen",
         "Artikel, Leiste aufklappen, \"Neu\"",
         "Neue Kategorie erscheint in der Liste und in der Kasse.",
         "RTH"),

        ("Kategorie ändern und löschen",
         "Artikel, Leiste aufklappen, \"Bearbeiten\"",
         "Umbenennen wirkt überall; Löschen fragt nach.",
         "RTH"),

        ("Neuen Artikel anlegen",
         "Artikel, \"+ Neuer Artikel\"",
         "Name, Preis, Kategorie reichen; der Artikel erscheint in der "
         "Kasse.",
         "RTH"),

        ("Stammdaten ändern",
         "Artikel, \"Bearbeiten\"",
         "Preis, Einkaufspreis, Kategorie, Sichtbarkeit lassen sich "
         "speichern.",
         "RTH"),

        ("Maske nutzt die ganze Breite",
         "Artikel, \"Bearbeiten\" (Telefon)",
         "Die Karten stehen untereinander, jede über die volle Breite; "
         "keine Zeile ist mittendrin abgeschnitten.",
         "H"),

        ("Preis und Bestand über den Nummernblock",
         "Artikel, Bearbeiten, ins Preisfeld tippen (Telefon)",
         "Die erste Ziffer ersetzt den Vorschlag; \"Bestätigen\" trägt "
         "den Wert ins Feld, Speichern behält ihn.",
         "H"),

        ("Artikel löschen",
         "Artikel, Kreuz in der Zeile",
         "Nach Rückfrage verschwindet der Artikel.",
         "RTH"),

        ("Bestellmenge erfassen",
         "Artikel, Mengenfeld",
         "Nummernblock öffnet sich, Menge wird übernommen.",
         "RTH"),

        ("Wareneingang buchen",
         "Artikel, \"Buchen\"",
         "Bestand steigt um die Menge, Einkaufspreis wird verrechnet.",
         "RTH"),

        ("Bestand korrigieren",
         "Artikel, Bearbeiten, Bestandskorrektur",
         "Der Bestand wird auf den gezählten Wert gesetzt, die "
         "Bewegung steht in der Historie.",
         "RTH"),

        ("Bestandshistorie",
         "Artikel, Bearbeiten",
         "Zugänge, Abgänge und Korrekturen mit Grund und Gerät.",
         "RTH"),

        ("Reihenfolge festlegen",
         "Artikel, \"Sortierung\"",
         "Die gewählte Reihenfolge gilt auch in der Kasse.",
         "RT"),

        ("Rezept anlegen",
         "Artikel, Mix-Artikel bearbeiten",
         "Zutaten zuordnen, Mengen und Einheiten ändern, entfernen.",
         "RTH"),

        ("Rezept: Zutatenfeld ist bedienbar",
         "Artikel, Mix-Artikel bearbeiten (Telefon)",
         "Zutat, Menge, Einheit und \"Hinzufügen\" stehen zweizeilig "
         "und sind alle zu treffen.",
         "H"),

        ("Zutat ohne eigenen Artikel",
         "Rezept, Freitextzutat",
         "Lässt sich mit Menge eintragen und wieder entfernen.",
         "RTH"),

        ("Verfügbarkeit und Kosten je Portion",
         "Artikel, Mix-Artikel",
         "\"Verfügbar\" richtet sich nach der knappsten Zutat; Kosten "
         "stimmen.",
         "RTH"),

        ("Flasche als Spirituose führen",
         "Artikel, Einheit \"Flasche\"",
         "Flaschengröße wird erfragt, Bestand läuft in ml.",
         "RTH"),

        ("Shot zur Flasche",
         "Artikel, Flasche bearbeiten",
         "Der Shot verkauft aus derselben Flasche; Bestand sinkt "
         "anteilig.",
         "RTH"),

        ("Einkaufsliste exportieren",
         "Artikel, \"Einkaufsliste exportieren\"",
         "CSV entsteht; darunter stehen Dateiname und Ordner.",
         "RTH"),

    ]),

    ("Kalender", [

        ("Monatsübersicht",
         "Events",
         "Der Monat wird angezeigt; Blättern vor und zurück geht.",
         "RTH"),

        ("Tag öffnen",
         "Events, Tag antippen",
         "Die Einträge des Tages erscheinen.",
         "RTH"),

        ("Eintrag anlegen",
         "Events, Tag, Neu",
         "Veranstaltung, Mitarbeiter oder Termin lassen sich anlegen.",
         "RTH"),

        ("Checkliste und Schichtplan mit anlegen",
         "Events, Veranstaltung anlegen",
         "Die beiden Häkchen legen Liste und Plan mit demselben Namen "
         "an.",
         "RTH"),

        ("Eintrag ändern und löschen",
         "Events, Eintrag antippen",
         "Änderungen werden gespeichert; Löschen fragt nach.",
         "RTH"),

        ("Zusammenspiel mit Kasse und Statistik",
         "Events / Kasse / Statistik",
         "Die Veranstaltung des Tages steht in der Kopfzeile und "
         "filtert die Statistik.",
         "RTH"),

    ]),

    ("Kassenbuch", [

        ("Zeitraum wählen",
         "Kassenbuch, Leiste unten \"Zeitraum\"",
         "Zugeklappt steht dort Monat und Jahr; aufgeklappt lassen "
         "sich beide wählen, die Tabelle zieht nach.",
         "RTH"),

        ("Zeile erfassen",
         "Kassenbuch, Formular",
         "Datum über den Kalender, Beträge über den Nummernblock; "
         "Speichern legt die Zeile an.",
         "RTH"),

        ("Startbestand wird vorbelegt",
         "Kassenbuch, neue Zeile",
         "Der Endbestand des Vortags steht als Startbestand.",
         "RTH"),

        ("Endbestand rechnet mit",
         "Kassenbuch, Beträge eintippen",
         "Der Endbestand ergibt sich, solange man ihn nicht selbst "
         "überschreibt.",
         "RTH"),

        ("Auffällige Zeilen",
         "Kassenbuch, Zeile mit falscher Rechnung",
         "Die Zeile wird als \"Prüfen\" hervorgehoben, mit Grund.",
         "RTH"),

        ("Zeile ändern und löschen",
         "Kassenbuch, Zeile antippen",
         "Änderungen werden übernommen; Löschen fragt nach.",
         "RTH"),

        ("Exportieren",
         "Kassenbuch, \"Excel exportieren\"",
         "Datei entsteht, zum Ausdrucken eingerichtet, mit "
         "Hinweisspalte.",
         "RTH"),

        ("Vier Spalten statt sieben",
         "Kassenbuch (Telefon)",
         "Datum, Einnahmen, Ausgaben, Endbestand - lesbar, nichts "
         "überlappt.",
         "H"),

    ]),

    ("Checkliste", [

        ("Liste anlegen und löschen",
         "Checkliste, \"Neue Liste\" / \"Löschen\"",
         "Liste erscheint; Löschen nimmt ihre Aufgaben mit.",
         "RTH"),

        ("Aufgabe eintragen",
         "Checkliste, Feld unten",
         "Die Aufgabe erscheint in der Liste.",
         "RTH"),

        ("Aufgabe abhaken",
         "Checkliste, Häkchen",
         "Der Fortschritt oben ändert sich mit.",
         "RTH"),

        ("Frist, Verantwortlich, Infos",
         "Checkliste, Felder der Zeile",
         "Alle Angaben werden gespeichert.",
         "RTH"),

        ("Aufgabe entfernen",
         "Checkliste, \"Entfernen\"",
         "Die Zeile verschwindet.",
         "RTH"),

        ("Exportieren",
         "Checkliste, \"Excel exportieren\"",
         "Datei entsteht; leere Liste gibt einen Hinweis statt einer "
         "leeren Datei.",
         "RTH"),

    ]),

    ("Schichtplan", [

        ("Plan anlegen und löschen",
         "Schichtplan, \"Plan anlegen\"",
         "Plan erscheint in der Liste.",
         "RTH"),

        ("Schicht eintragen",
         "Schichtplan, Feld unten",
         "Tätigkeit, Zeiten und Bedarf lassen sich eintragen.",
         "RTH"),

        ("Helfer eintragen",
         "Schichtplan, Helferfeld",
         "Die Zahl \"eingetragen\" steigt.",
         "RTH"),

        ("Farben stimmen",
         "Schichtplan",
         "Grün besetzt, orange teilweise, rot niemand.",
         "RTH"),

        ("Schichten übernehmen",
         "Schichtplan, \"Schichten übernehmen\"",
         "Die Schichten eines anderen Plans werden kopiert.",
         "RTH"),

        ("Exportieren",
         "Schichtplan, \"Excel exportieren\"",
         "Datei entsteht; darunter stehen Name und Ordner.",
         "RTH"),

    ]),

    ("Statistik", [

        ("Verkaufsliste",
         "Statistik",
         "Verkäufe des Zeitraums stehen in der Tabelle.",
         "RTH"),

        ("Nach Event filtern",
         "Statistik, Leiste unten \"Auswahl\"",
         "Nur Verkäufe dieser Veranstaltung; die zugeklappte Leiste "
         "nennt das gewählte Event.",
         "RTH"),

        ("Nach Zeitraum filtern",
         "Statistik, Leiste unten, Von / Bis",
         "Kalender öffnet sich, Filter wirkt, Kreuz setzt zurück; die "
         "Leiste nennt den Zeitraum.",
         "RTH"),

        ("Einzelne Position löschen",
         "Statistik, Zeile wählen, \"Ausgewählte löschen\"",
         "Nach Rückfrage weg; Tagesumsatz und Bestand ziehen nach.",
         "RTH"),

        ("Ganzen Zeitraum löschen",
         "Statistik, \"Zeitraum löschen\"",
         "Nach deutlicher Rückfrage sind die Verkäufe des Zeitraums "
         "weg.",
         "RTH"),

        ("Gesamtverkaufszahlen",
         "Statistik, Karte unten",
         "Einnahmen, Ausgaben, Gewinn passen zum Filter.",
         "RTH"),

        ("Top-Artikel",
         "Statistik, Karte unten",
         "Rangliste passt zum Filter.",
         "RTH"),

        ("Verteilung nach Kategorie",
         "Statistik, Tortendiagramm",
         "Diagramm und Legende passen zum Filter.",
         "RT"),

        ("Einkaufspreise nachtragen",
         "Statistik, Hinweiszeile",
         "Fehlende Rezeptpreise lassen sich nachtragen; Gewinn "
         "stimmt danach.",
         "RTH"),

        ("Exportieren",
         "Statistik, \"Excel exportieren\"",
         "Datei entsteht; darunter stehen Name und Ordner.",
         "RTH"),

    ]),

    ("Einstellungen", [

        ("Farbmodus hell / dunkel",
         "Einstellungen, Farbmodus",
         "Die Oberfläche wechselt vollständig; nach Neustart bleibt "
         "die Wahl.",
         "RTH"),

        ("Hoch- oder Querformat",
         "Einstellungen, Bildschirmausrichtung",
         "Die Anordnung wechselt; nach Neustart bleibt die Wahl.",
         "RT"),

        ("Kein Querformat auf dem Telefon",
         "Einstellungen (Telefon)",
         "Die Auswahl fehlt, und das Gerät bleibt beim Drehen im "
         "Hochformat.",
         "H"),

        ("Demo-Modus starten",
         "Einstellungen, \"Demo starten\"",
         "Akzentfarbe wird grün, oben steht DEMO; alles lässt sich "
         "ausprobieren.",
         "RTH"),

        ("Demo-Modus beenden",
         "Einstellungen, \"Demo beenden\"",
         "Alles Ausprobierte ist verworfen, der Stand von vorher gilt "
         "wieder.",
         "RTH"),

        ("Gerät umbenennen",
         "Einstellungen, \"Gerät umbenennen\"",
         "Der neue Name steht in der Kopfzeile und in den Übergaben.",
         "RTH"),

        ("Übergaben anzeigen",
         "Einstellungen, \"Übergaben anzeigen\"",
         "Das Protokoll zeigt, wer wann an wen übergeben hat.",
         "RTH"),

    ]),

    ("Mehrere Geräte", [

        ("Daten empfangen: warten",
         "Einstellungen, \"Daten empfangen\"",
         "Das Gerät wartet und zeigt Name und Adresse.",
         "RTH"),

        ("Daten senden: suchen",
         "Einstellungen, \"Daten senden\"",
         "Das wartende Gerät steht in der Liste.",
         "RTH"),

        ("Gegenseite wird gefragt",
         "Senden, Gerät antippen",
         "Auf dem anderen Gerät erscheint \"möchte Daten senden - "
         "annehmen?\".",
         "RTH"),

        ("Ablehnen",
         "Empfangen, \"Ablehnen\"",
         "Der Sender meldet die Ablehnung, nichts wird übertragen.",
         "RTH"),

        ("Datenbank übertragen",
         "Senden, \"Datenbank\"",
         "Das andere Gerät hat danach alle Artikel und ist "
         "Nebengerät.",
         "RTH"),

        ("Kasse übertragen",
         "Senden, \"Kasse\"",
         "Das Schreibrecht wandert; dieses Gerät ist danach nur noch "
         "Ansicht.",
         "RTH"),

        ("Buchungen übertragen",
         "Senden, \"Buchungen\"",
         "Nur Zugänge; zweimal gesendet ändert nichts.",
         "RTH"),

        ("Kasse steht auf dem Nebengerät nicht zur Wahl",
         "Senden auf einem Nebengerät",
         "Nur Datenbank und Buchungen werden angeboten, mit Begründung.",
         "RTH"),

        ("Ohne WLAN: Datei schreiben",
         "Senden, \"Stattdessen als Datei\"",
         "Datei entsteht; danach lässt sie sich teilen.",
         "RTH"),

        ("Ohne WLAN: Datei einlesen",
         "Empfangen, \"Kein Netz? Datei suchen\"",
         "Die Datei wird gefunden und nach Rückfrage eingespielt.",
         "RTH"),

        ("Empfangene Datei wird gefunden",
         "Empfangen, Dateiliste",
         "Auch Dateien aus Downloads, Desktop oder Bluetooth stehen "
         "dort, mit Herkunft.",
         "R"),

        ("Dateiauswahl des Geräts",
         "Empfangen, \"Datei suchen\"",
         "Androids Dateiauswahl öffnet sich; die gewählte Datei wird "
         "übernommen.",
         "TH"),

        ("Per Bluetooth senden",
         "Senden, Datei, \"Per Bluetooth senden\"",
         "Die Bluetooth-Übertragung öffnet sich; die Gegenseite nimmt "
         "an.",
         "TH"),

        ("Nebengerät: Sperren sichtbar",
         "Artikel auf einem Nebengerät",
         "Neu, Sortierung, Buchen, Bearbeiten und Löschen sind grau, "
         "darüber steht der Grund.",
         "RTH"),

        ("Nebengerät: kein Absturz",
         "Nebengerät, gesperrte Knöpfe antippen",
         "Nichts passiert bzw. ein wegtippbarer Hinweis - das Programm "
         "läuft weiter.",
         "RTH"),

        ("Nebengerät: buchen bleibt erlaubt",
         "Nebengerät, Kasse und Listen",
         "Verkaufen, Kassenbuch, Checklisten und Schichten gehen.",
         "RTH"),

        ("Bonnummern je Gerät",
         "Zwei Geräte verkaufen",
         "Keine doppelten Bonnummern nach dem Einsammeln.",
         "RTH"),

        ("Bestand über mehrere Geräte",
         "Zwei Geräte verkaufen, dann einsammeln",
         "Die Abgänge beider Geräte sind zusammengerechnet.",
         "RTH"),

    ]),

    ("Ausgabe und Teilen", [

        ("Ordner steht beim Export dabei",
         "Nach jedem Export",
         "Dateiname und Ordner stehen unter dem Knopf.",
         "RTH"),

        ("Teilen: Statistik",
         "Statistik, \"Teilen\"",
         "Teilen-Auswahl (Android) bzw. Ordner mit ausgewählter Datei "
         "(Rechner).",
         "RTH"),

        ("Teilen: Kassenbuch",
         "Kassenbuch, \"Teilen\"",
         "wie oben",
         "RTH"),

        ("Teilen: Checkliste",
         "Checkliste, \"Teilen\"",
         "wie oben",
         "RTH"),

        ("Teilen: Schichtplan",
         "Schichtplan, \"Teilen\"",
         "wie oben",
         "RTH"),

        ("Teilen: Einkaufsliste",
         "Artikel, \"Teilen\"",
         "wie oben",
         "RTH"),

        ("Teilen ohne Export",
         "Teilen antippen, bevor exportiert wurde",
         "Hinweis \"Erst exportieren, dann teilen.\"",
         "RTH"),

        ("Geteilte Datei ist auffindbar",
         "Android, nach dem Teilen",
         "Die Datei liegt in Download/KiG POS und ist im Dateimanager "
         "zu sehen.",
         "TH"),

        ("Handbuch als PDF",
         "Handbuch, \"Als PDF exportieren\"",
         "PDF mit allen Bildern entsteht.",
         "R"),

        ("PDF-Export gesperrt",
         "Handbuch auf Android",
         "Der Knopf ist gesperrt und sagt, warum.",
         "TH"),

    ]),

    ("Handbuch", [

        ("Themen und Anleitung",
         "Userguide",
         "Links das Thema wählen, rechts erscheint der Text mit Bild.",
         "RTH"),

        ("Alle Themen öffnen sich",
         "Userguide, jedes Thema",
         "Kein Thema bleibt leer oder bricht ab.",
         "RTH"),

        ("Bilder passen zum Text",
         "Userguide",
         "Die Screenshots zeigen den beschriebenen Bildschirm.",
         "RTH"),

    ]),

    ("Darstellung", [

        ("Nichts ragt über den Rand",
         "Alle Bildschirme",
         "Keine Schaltfläche und keine Beschriftung läuft aus dem Bild.",
         "RTH"),

        ("Keine abgeschnittenen Wörter",
         "Alle Bildschirme",
         "Nichts bricht mitten im Wort um.",
         "RTH"),

        ("Kopfzeile schlank",
         "Alle Bildschirme (Telefon)",
         "Die Kopfzeile nimmt deutlich weniger als ein Zehntel der "
         "Höhe ein; eine Fußzeile gibt es nicht mehr.",
         "H"),

        ("Filterleiste sagt, was gilt",
         "Artikel, Kassenbuch, Statistik",
         "Zugeklappt steht dort der eingestellte Stand - ohne dass man "
         "sie öffnen muss.",
         "RTH"),

        ("Filterleiste klappt auf und zu",
         "Artikel, Kassenbuch, Statistik",
         "Ein Tipp öffnet sie, ein zweiter schließt sie; beim Wechsel "
         "des Bildschirms ist sie wieder zu.",
         "RTH"),

        ("Bildschirmtastatur schiebt das Feld hoch",
         "Listen, neues Feld beschreiben",
         "Das beschriebene Feld bleibt sichtbar.",
         "TH"),

        ("Nummernblock ohne Systemtastatur",
         "Zahlenfelder antippen",
         "Nur der Nummernblock geht auf, nicht zusätzlich die "
         "Tastatur.",
         "TH"),

        ("Dunkelmodus überall",
         "Alle Bildschirme im Dunkelmodus",
         "Kein weißer Kasten, keine unlesbare Schrift.",
         "RTH"),

        ("Lesbarkeit auf E-Ink",
         "Alle Bildschirme auf dem Tablet",
         "Kontraste reichen, nichts verschwimmt.",
         "T"),

    ]),

]


# =========================================================
# Mappe bauen
# =========================================================

def _kopfzeile(blatt, spalten, zeile=1):

    for nummer, (titel, breite) in enumerate(spalten, start=1):

        zelle = blatt.cell(row=zeile, column=nummer, value=titel)

        zelle.font = Font(bold=True, color="FFFFFF", size=11)
        zelle.fill = PatternFill("solid", fgColor=FARBEN["kopf"])
        zelle.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        blatt.column_dimensions[get_column_letter(nummer)].width = breite

    blatt.row_dimensions[zeile].height = 28


def _rahmen():

    duenn = Side(style="thin", color="D0D0D0")

    return Border(left=duenn, right=duenn, top=duenn, bottom=duenn)


def _anleitung(mappe):

    blatt = mappe.create_sheet("Anleitung")

    blatt.column_dimensions["A"].width = 26
    blatt.column_dimensions["B"].width = 92

    zeilen = [
        ("KiG POS - Prüfkatalog", ""),
        ("", ""),
        ("Wozu", "Jede Funktion einmal bewusst ausprobieren und "
                 "festhalten, was dabei herauskam."),
        ("Aufbau", "Ein Blatt je Bereich. Je Zeile eine Funktion, je "
                   "Gerät eine Spalte."),
        ("", ""),
        ("Geräte", "Rechner = Windows, Tablet = Boox Go 10.3, "
                   "Handy = Telefon im Hochformat."),
        ("Strich (entfällt)", "Die Funktion gibt es auf diesem Gerät "
                              "nicht - dort ist nichts zu prüfen."),
        ("", ""),
        ("Zustände", "offen - noch nicht geprüft"),
        ("", "OK - tut, was danebensteht"),
        ("", "Fehler - tut es nicht"),
        ("", "Teilweise - tut es, aber nicht ganz"),
        ("", "entfällt - kommt auf diesem Gerät nicht vor"),
        ("", ""),
        ("Kommentar", "Was aufgefallen ist. Bei \"Fehler\" bitte so "
                      "genau wie möglich: Was hast du getan, was "
                      "passierte, was hattest du erwartet?"),
        ("", ""),
        ("Übersicht", "Das zweite Blatt zählt automatisch zusammen, "
                      "wie weit du bist."),
        ("", ""),
        ("Neu erzeugen", "werkzeuge/pruefkatalog.py - erzeugt diese "
                         "Mappe neu, wenn Funktionen dazukommen. "
                         "Achtung: Kommentare gehen dabei verloren, "
                         "also vorher sichern."),
    ]

    for nummer, (links, rechts) in enumerate(zeilen, start=1):

        a = blatt.cell(row=nummer, column=1, value=links)
        b = blatt.cell(row=nummer, column=2, value=rechts)

        a.font = Font(bold=True, size=14 if nummer == 1 else 11,
                      color=FARBEN["kopf"] if nummer == 1 else "000000")
        b.alignment = Alignment(vertical="top", wrap_text=True)

    return blatt


def _altbestand(ziel):
    """Liest Ergebnisse und Kommentare aus einer vorhandenen Mappe.

    Ergebnis: {(Bereich, Funktion): ([Rechner, Tablet, Handy],
    Kommentar)}. Fehlt die Datei oder laesst sie sich nicht lesen,
    bleibt das Ergebnis leer - dann entsteht eben eine frische Mappe.
    """

    ziel = Path(ziel)

    if not ziel.exists():
        return {}

    try:
        from openpyxl import load_workbook

        mappe = load_workbook(ziel, data_only=True)

    except Exception as fehler:                      # pragma: no cover
        print(f"   (vorhandene Mappe nicht lesbar: {fehler})")
        return {}

    bestand = {}

    for name in mappe.sheetnames:

        if name in ("Anleitung", "Übersicht"):
            continue

        blatt = mappe[name]

        for zeile in blatt.iter_rows(min_row=2):

            werte = [zelle.value for zelle in zeile]

            if len(werte) < 8 or not werte[1]:
                continue

            funktion = str(werte[1]).strip()

            zustaende = [
                str(wert).strip() if wert else ""
                for wert in werte[4:7]
            ]

            kommentar = str(werte[7]).strip() if werte[7] else ""

            bestand[(name, funktion)] = (zustaende, kommentar)

    mappe.close()

    return bestand


def _uebernehmen(bestand, bereich, funktion):
    """Was fuer diese Funktion schon eingetragen war.

    Liefert (Zustaende, Kommentar). Zustaende ist None, wenn die
    Funktion sich geaendert hat und deshalb neu angesehen werden will.
    """

    schluessel = (bereich, funktion)

    # Alte Namen mitnehmen
    for alt, neu in UMBENANNT.items():

        if neu == schluessel and alt in bestand:

            _zustaende, kommentar = bestand[alt]

            return None, kommentar

    if schluessel not in bestand:
        return None, ""

    zustaende, kommentar = bestand[schluessel]

    if schluessel in GEAENDERT:
        return None, kommentar

    return zustaende, kommentar


def _bereichsblatt(mappe, bereich, eintraege, pruefliste, bestand):

    blatt = mappe.create_sheet(bereich[:31])

    spalten = [
        ("Nr", 5),
        ("Funktion", 34),
        ("Wo zu finden", 34),
        ("Erwartetes Verhalten", 52),
    ]

    spalten += [(geraet, 12) for geraet in GERAETE]
    spalten += [("Kommentar", 46)]

    _kopfzeile(blatt, spalten)

    rahmen = _rahmen()

    for nummer, (funktion, wo, erwartet, geraete) in enumerate(
            eintraege, start=1
    ):

        zeile = nummer + 1

        werte = [nummer, funktion, wo, erwartet]

        alte_zustaende, alter_kommentar = _uebernehmen(
            bestand, bereich, funktion
        )

        for stelle, (kuerzel, geraet) in enumerate(zip("RTH", GERAETE)):

            if kuerzel not in geraete:
                werte.append(NICHT)
                continue

            frueher = (
                alte_zustaende[stelle] if alte_zustaende else ""
            )

            # "entfällt" von frueher gilt nicht mehr, wenn es die
            # Funktion auf dem Geraet inzwischen gibt.
            if frueher in ("", NICHT):
                werte.append("offen")
            else:
                werte.append(frueher)

        werte.append(alter_kommentar)

        for spalte, wert in enumerate(werte, start=1):

            zelle = blatt.cell(row=zeile, column=spalte, value=wert)

            zelle.border = rahmen
            zelle.alignment = Alignment(
                vertical="top",
                wrap_text=spalte in (2, 3, 4, 8),
                horizontal="center" if spalte in (1, 5, 6, 7) else "left",
            )

        blatt.row_dimensions[zeile].height = 30

        pruefliste.append((bereich, funktion, geraete))

    letzte = len(eintraege) + 1

    # Auswahlliste in den Gerätespalten
    pruefung = DataValidation(
        type="list",
        formula1='"' + ",".join(ZUSTAENDE) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    blatt.add_data_validation(pruefung)

    for spalte in range(5, 5 + len(GERAETE)):

        buchstabe = get_column_letter(spalte)

        pruefung.add(f"{buchstabe}2:{buchstabe}{letzte}")

        bereich_text = f"{buchstabe}2:{buchstabe}{letzte}"

        for zustand, farbe in (
            ("OK", FARBEN["ok"]),
            ("Fehler", FARBEN["fehler"]),
            ("Teilweise", FARBEN["teilweise"]),
            (NICHT, FARBEN["entfaellt"]),
        ):
            blatt.conditional_formatting.add(
                bereich_text,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{zustand}"'],
                    fill=PatternFill("solid", fgColor=farbe),
                ),
            )

    blatt.freeze_panes = "B2"
    blatt.auto_filter.ref = f"A1:H{letzte}"

    return blatt


def _uebersicht(mappe, bereiche):
    """Zählt je Bereich und Gerät zusammen - mit Formeln, damit sich
    die Zahlen beim Ausfüllen mitbewegen."""

    blatt = mappe.create_sheet("Übersicht", 1)

    spalten = [("Bereich", 26), ("Funktionen", 12)]

    for geraet in GERAETE:
        spalten += [
            (f"{geraet}: OK", 13),
            (f"{geraet}: Fehler", 14),
            (f"{geraet}: offen", 13),
        ]

    _kopfzeile(blatt, spalten)

    rahmen = _rahmen()

    for nummer, (bereich, anzahl) in enumerate(bereiche, start=1):

        zeile = nummer + 1

        blattname = f"'{bereich[:31]}'"

        werte = [bereich, anzahl]

        for spalte in ("E", "F", "G"):

            bezug = f"{blattname}!{spalte}2:{spalte}{anzahl + 1}"

            werte += [
                f'=COUNTIF({bezug},"OK")',
                f'=COUNTIF({bezug},"Fehler")',
                f'=COUNTIF({bezug},"offen")',
            ]

        for spalte, wert in enumerate(werte, start=1):

            zelle = blatt.cell(row=zeile, column=spalte, value=wert)
            zelle.border = rahmen
            zelle.alignment = Alignment(
                horizontal="left" if spalte == 1 else "center"
            )

    # Summenzeile
    summe = len(bereiche) + 2

    blatt.cell(row=summe, column=1, value="Zusammen").font = Font(bold=True)

    for spalte in range(2, 2 + 1 + 3 * len(GERAETE)):

        buchstabe = get_column_letter(spalte)

        zelle = blatt.cell(
            row=summe, column=spalte,
            value=f"=SUM({buchstabe}2:{buchstabe}{len(bereiche) + 1})",
        )
        zelle.font = Font(bold=True)
        zelle.fill = PatternFill("solid", fgColor=FARBEN["bereich"])
        zelle.alignment = Alignment(horizontal="center")

    blatt.freeze_panes = "B2"

    return blatt


def erzeugen(ziel):

    # Was schon geprueft und kommentiert wurde, bleibt erhalten.
    bestand = _altbestand(ziel)

    mappe = Workbook()
    mappe.remove(mappe.active)

    _anleitung(mappe)

    pruefliste = []
    bereiche = []

    for bereich, eintraege in KATALOG:

        _bereichsblatt(mappe, bereich, eintraege, pruefliste, bestand)

        bereiche.append((bereich, len(eintraege)))

    _uebersicht(mappe, bereiche)

    ziel = Path(ziel)
    ziel.parent.mkdir(parents=True, exist_ok=True)

    try:
        mappe.save(ziel)

    except PermissionError:

        # Excel haelt die Datei offen. Statt die Arbeit wegzuwerfen,
        # daneben ablegen - der Name sagt, was zu tun ist.
        ausweich = ziel.with_name(f"{ziel.stem}_NEU{ziel.suffix}")

        mappe.save(ausweich)

        print(f"   {ziel.name} ist geöffnet - neu geschrieben nach "
              f"{ausweich.name}")

        ziel = ausweich

    return ziel, pruefliste, bestand


def main():

    if len(sys.argv) > 1:
        ziel = Path(sys.argv[1])
    else:
        ziel = config.EXPORT_EXCEL_DIR / "KiG_POS_Pruefkatalog.xlsx"

    ziel, pruefliste, bestand = erzeugen(ziel)

    gesamt = len(pruefliste)

    einzelpruefungen = sum(len(g) for _b, _f, g in pruefliste)

    print(f"{ziel}")
    print(f"{len(KATALOG)} Bereiche, {gesamt} Funktionen, "
          f"{einzelpruefungen} Einzelprüfungen")

    if bestand:

        namen = {(b, f) for b, f, _g in pruefliste}

        wieder = len(namen & set(bestand))
        neu = len(namen - set(bestand))
        weg = len(set(bestand) - namen)

        print(f"Aus der bisherigen Mappe übernommen: {wieder} Funktionen, "
              f"{neu} neu, {weg} entfallen")

    for bereich, eintraege in KATALOG:
        print(f"   {bereich:22s} {len(eintraege):3d}")


if __name__ == "__main__":
    main()
